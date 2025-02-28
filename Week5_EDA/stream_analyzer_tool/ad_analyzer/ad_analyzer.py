import json
import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import re
import argparse
import subprocess
from urllib.parse import urlparse

# Get the parent directory of the current file
parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
# Add the parent directory to sys.path
sys.path.insert(0, parent_dir)
import config.project_config as config
from utils.stream_info import get_stream_info, extract_audio_video_pids
from utils.frame_info import (
    extract_frame_information,
    get_pts_duration_from_packet_info,
    get_all_packets,
)
from utils.media_info import extract_media_info

global output_dir
output_dir = os.path.join(os.path.dirname(__file__), "..", "output_data")


def load_standard_values(file_path):
    file_path = os.path.join(parent_dir, "config", file_path)
    with open(file_path, "r") as file:
        return json.load(file)


# Load project specific configs
standard_values = load_standard_values(config.project_spec_json[config.project_name])
keys_not_expected = standard_values.get("keys_not_expected", {})
expected_audio_track_count = standard_values.get("expected_audio_track_count", 0)
expected_video_track_count = standard_values.get("expected_video_track_count", 0)
supported_audio_formats = standard_values.get("supported_audio_formats", [])
supported_ac3_channels = standard_values.get("supported_ac3_channels", [])
order_of_audio_tracks = standard_values.get("order_of_audio_tracks", {})
global final_excel_path


def is_url(path):
    parsed_url = urlparse(path)
    return parsed_url.scheme in ("http", "https")


def generate_remarks(data, standard_values, key_prefix, keys_not_expected=None):
    """
    Generates remarks for data based on standard values and keys not expected.
    """
    return [
        (
            "Ok"
            if k in standard_values and str(v) == str(standard_values.get(k))
            else (
                "Ok"
                if k not in standard_values
                and v is not None
                and (not keys_not_expected or k not in keys_not_expected[key_prefix])
                else (
                    "Property is not present"
                    if v is None
                    else (
                        "Property not expected for this format"
                        if keys_not_expected and k in keys_not_expected[key_prefix]
                        else "Property value is not as per standard. Expected is {}".format(
                            str(standard_values.get(k))
                        )
                    )
                )
            )
        )
        for k, v in data.items()
    ]


def generate_std_values(data, standard_values):
    """
    Generates standard values for data based on standard values from project config.
    """
    return [standard_values.get(k, str(data.get(k, 0))) for k in data.keys()]


def process_media_data(
    media_data,
    standard_values,
    keys_not_expected=None,
    supported_formats=None,
    supported_channels=None,
):
    """
    Processes media data (video or audio) and writes it to a DataFrame.
    """
    media_json = []
    media_df = []
    remarks_column_data = []
    std_column_data = []

    for media in media_data:
        format = media.get("Format")
        channels = media.get("Channel(s)", "")

        if supported_formats and format not in supported_formats:
            continue
        if (
            supported_channels
            and format == "AC-3"
            and channels not in supported_channels
        ):
            continue

        key_prefix = None
        if format == "AVC":
            key_prefix = "AVC_Video"
        elif format == "AC-3" and channels == "2 channels":
            key_prefix = "AC3_2_channels"
        elif format == "AC-3" and channels == "6 channels":
            key_prefix = "AC3_6_channels"
        elif format == "MPEG Audio":
            key_prefix = "MPEG_Audio"
        elif format == "AAC LC SBR" and channels == "2 channels":
            key_prefix = "AAC_LC_SBR_2_channels"
        elif format == "E-AC-3" and channels == "6 channels":
            key_prefix = "EAC3_6_channels"

        if key_prefix:
            remarks = generate_remarks(
                media, standard_values[key_prefix], key_prefix, keys_not_expected
            )
            std_values = generate_std_values(media, standard_values[key_prefix])

            remarks_column_data.append(remarks)
            std_column_data.append(std_values)

            media_json.append(flatten_json(media))
            data = list(media_json[-1].items())

            for i in range(len(std_column_data[0])):
                if remarks_column_data[0][i].startswith("Property not expected"):
                    std_column_data[0][i] = " "

            updated_data_tuples = [
                (k, v, v1, remark)
                for (k, v), v1, remark in zip(data, std_values, remarks)
            ]
            media_df.append(
                pd.DataFrame(
                    updated_data_tuples,
                    columns=["Item", "Actual Value", "Expected Value", "Remarks"],
                )
            )

    return media_df


def get_excel_path():
    return output_dir


def write_data_to_excel(
    output_path, video_df, audio_df, audio_analysis_df, video_analysis_df
):
    """
    Writes data to an Excel file and formats it.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for i, df in enumerate(video_df):
            df.to_excel(writer, sheet_name=f"Media Info Video Track {i+1}", index=False)
        for i, df in enumerate(audio_df):
            df.to_excel(writer, sheet_name=f"Media Info Audio Track {i+1}", index=False)
        for i, df in enumerate(audio_analysis_df):
            df.to_excel(writer, sheet_name=f"Audio Analysis Track {i+1}", index=False)
        for i, df in enumerate(video_analysis_df):
            df.to_excel(writer, sheet_name=f"Video Analysis Track {i+1}", index=False)


def main():
    parser = argparse.ArgumentParser(
        description="Process a .ts file from a network path."
    )

    # Define the common arguments
    parser.add_argument(
        "file_paths",
        nargs="*",
        type=str,
        help="Path(s) to the .ts file(s) or directory containing .ts files.",
    )
    parser.add_argument(
        "--run-check-services",
        action="store_true",
        help="Run checkServices.py before processing .ts files.",
    )
    parser.add_argument(
        "--check-services-args",
        nargs=argparse.REMAINDER,
        help="Arguments to pass to checkServices.py.",
    )

    args = parser.parse_args()

    if args.run_check_services:
        if not args.check_services_args:
            print(
                "No arguments provided for checkServices.py. Please provide arguments like -c <cardID> -ep <endpoint> -s"
            )
            sys.exit(1)

        try:
            result = subprocess.run(
                ["python", "checkServices.py"] + args.check_services_args,
                capture_output=True,
                check=True,
                text=True,
            )
        except subprocess.CalledProcessError as e:
            print(f"Error while running checkServices.py: {e}")
            print(f"Error: {e.stderr}")
            sys.exit(1)

        # Extract the folder name from the output
        folder_name = result.stdout.strip()

        if not folder_name:
            print("Error: No folder name returned by checkServices.py.")
            sys.exit(1)

        print(f"TS files will be processed from folder: {folder_name}")

        # Process all TS files in the folder
        ts_files = [
            os.path.join(folder_name, f)
            for f in os.listdir(folder_name)
            if f.endswith(".ts")
        ]

        if not ts_files:
            print("No .ts files found in the specified folder.")
            sys.exit(1)

    elif args.file_paths:

        input_path = args.file_paths[0]

        # Determine if input_path is a file or directory
        if is_url(input_path):  # It's a URL, process it directly
            ts_files = args.file_paths
        elif os.path.isfile(input_path):
            ts_files = args.file_paths
        elif os.path.isdir(input_path):
            ts_files = [
                os.path.join(input_path, f)
                for f in os.listdir(input_path)
                if f.endswith(".ts")
            ]
        else:
            print(f"Invalid path provided: '{input_path}'")
            sys.exit(1)
    else:
        print(
            "Example, for single ts file processing : python ad_analyzer.py osn-demo-1_BXP-P0144012.ts \n"
        )
        print(
            "Example, to process multiple ts files in folder : python ad_analyzer.py ts_files_folder \n"
        )
        print(
            "Example, to run check-services with card id 645 and prod server and then do processing :"
        )
        print(
            "\tpy ad_analyzer.py --run-check-services --check-services-args -c 645 -ep prod -s"
        )
        print(
            """usage: checkServices.py [-h] [-ep [{prod,int,demo,eu5}]] -c CARDID
                        [-d DEVICEID] [-s]
                        [-l {DEBUG,INFO,WARNING,ERROR,CRITICAL}]
                checkServices.py: error: the following arguments are required: -c/--cardID"""
        )
        sys.exit(1)

    for ts_file in ts_files:
        ts_file_name = os.path.splitext(os.path.basename(ts_file))[0]

        output_dir = os.path.join(
            os.path.dirname(__file__), "output_data", ts_file_name
        )
        os.makedirs(output_dir, exist_ok=True)

        output_dict = {"Ad_media_info": {}, "Ad_Analysis": {"Audio": [], "Video": []}}

        # Process stream info
        stream_info = get_stream_info(ts_file)
        dump_json(
            stream_info, os.path.join(output_dir, f"streams_data_{ts_file_name}.json")
        )

        media_info_av_data = extract_media_info(ts_file)
        output_dict["Ad_media_info"] = media_info_av_data

        audio_pids, video_pids = extract_audio_video_pids(stream_info)
        audio_charts = []
        video_charts = []

        audio_pts, video_pts = process_streams(
            ts_file,
            audio_pids,
            video_pids,
            ts_file_name,
            output_dir,
            output_dict,
            audio_charts,
            video_charts,
            stream_info,
            media_info_av_data,
        )
        check_audio_boundaries(audio_pts, video_pts, output_dict)

        # Write final data to JSON
        dump_json(
            output_dict, os.path.join(output_dir, f"media_info_{ts_file_name}.json")
        )

        # Write data to Excel
        video_data = output_dict.get("Ad_media_info", {}).get("Video", [])
        audio_data = output_dict.get("Ad_media_info", {}).get("Audio", [])
        audio_analysis_data = output_dict.get("Ad_Analysis", {}).get("Audio", [])
        video_analysis_data = output_dict.get("Ad_Analysis", {}).get("Video", [])

        video_df, audio_df = [], []

        video_df = process_media_data(video_data, standard_values)
        audio_df = process_media_data(
            audio_data,
            standard_values,
            keys_not_expected,
            supported_audio_formats,
            supported_ac3_channels,
        )
        audio_analysis_df = [
            pd.DataFrame(list(flatten_json(analysis).items()), columns=["Key", "Value"])
            for analysis in audio_analysis_data
        ]
        video_analysis_df = [
            pd.DataFrame(list(flatten_json(analysis).items()), columns=["Key", "Value"])
            for analysis in video_analysis_data
        ]

        output_excel_path = f"media_info_{ts_file_name}.xlsx"
        output_path = os.path.join(output_dir, output_excel_path)
        write_data_to_excel(
            output_path, video_df, audio_df, audio_analysis_df, video_analysis_df
        )

        # Add a "Remarks" column to specific sheets
        add_remarks_column_to_specific_sheets(
            output_path, "Audio Analysis Track ", remark_generator, media_info_av_data
        )
        add_remarks_column_to_specific_sheets(
            output_path, "Video Analysis Track ", remark_generator, media_info_av_data
        )

        # Add charts to excel
        insert_charts_into_excel(
            output_path, video_charts, audio_charts, img_width=800, img_height=400
        )

        workbook = load_workbook(output_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            set_column_widths(sheet)
        workbook.save(output_path)

        filter_remarks(
            output_path,
            audio_data,
            video_data,
            supported_audio_formats,
            supported_ac3_channels,
            expected_video_track_count,
            expected_audio_track_count,
        )
        # Load the workbook
        workbook = load_workbook(output_path)
        # Set column widths for all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            set_column_widths(sheet)

        # Save the workbook
        workbook.save(output_path)

        final_excel_path = output_path
        print(f"Flattened JSON data written to {output_path}")


def dump_json(data, path):
    with open(path, "w") as json_file:
        json.dump(data, json_file, indent=4)


def plot_pkt_sizes_and_bitrates(
    frames, bitrates, codec_type, stream_id, dict, ts_file_name, output_dir
):
    """
    Calculates max,min bitrates and plots a graph of bitrate and frame size against time.
    """

    bitrate_times = [br[0] for br in bitrates]
    bitrate_values = [br[1] / 1000 for br in bitrates]

    max_bit_rate = max(bitrate_values)
    min_bit_rate = min(bitrate_values)
    avg_bit_rate = sum(bitrate_values) / len(bitrate_values)

    max_bit_rate_time = bitrate_times[bitrate_values.index(max_bit_rate)]
    min_bit_rate_time = bitrate_times[bitrate_values.index(min_bit_rate)]

    dict["Max Bitrate"] = str(f"{max_bit_rate:.3f}") + " kbps"
    dict["Min Bitrate"] = str(f"{min_bit_rate:.3f}") + " kbps"
    dict["Average Bitrate"] = str(f"{avg_bit_rate:.3f}") + " kbps"
    dict["Time at Max Bitrate"] = str(f"{max_bit_rate_time:.3f}") + " s"
    dict["Time at Min Bitrate"] = str(f"{min_bit_rate_time:.3f}") + " s"

    packet_sizes = []
    packet_times = []

    for packet in frames:
        packet_sizes.append(int(packet["pkt_size"]) / 1024)
        packet_times.append(float(packet["pts_time"]))

    max_packet_size = max(packet_sizes)
    min_packet_size = min(packet_sizes)
    avg_packet_size = sum(packet_sizes) / len(packet_sizes)

    max_packet_size_time = packet_times[packet_sizes.index(max_packet_size)]
    min_packet_size_time = packet_times[packet_sizes.index(min_packet_size)]

    dict["Max Frame size"] = str(f"{max_packet_size:.3f}") + " KB"
    dict["Min Frame size"] = str(f"{min_packet_size:.3f}") + " KB"
    dict["Average Frame size"] = str(f"{avg_packet_size:.3f}") + " KB"
    dict["Time at Max Frame size"] = str(f"{max_packet_size_time:.3f}") + " s"
    dict["Time at Min Frame size"] = str(f"{min_packet_size_time:.3f}") + " s"

    # Plot packet sizes
    plt.figure(figsize=(12, 6))  # width,height
    plt.suptitle(
        "Packet Sizes and Bitrate Over Time For " + codec_type + stream_id,
        fontsize=16,
        fontweight="bold",
    )
    plt.subplot(2, 1, 1)
    # x,y
    plt.plot(packet_times, packet_sizes, label="Frame Size")
    plt.xticks(
        range(int(packet_times[0]), int(packet_times[-1]) + 1, 1)
    )  # Set x-ticks to 1-second intervals
    plt.xlabel("Time (s)")
    plt.ylabel("Frame Size (KB)")
    plt.title("Frame Sizes Over Time")
    plt.legend()

    # Plot bitrate
    plt.subplot(2, 1, 2)
    plt.plot(bitrate_times, bitrate_values, label="Bitrate", color="orange")
    plt.xticks(
        range(int(bitrate_times[0]), int(bitrate_times[-1]) + 1, 1)
    )  # Set x-ticks to 1-second intervals
    plt.xlabel("Time (s)")
    plt.ylabel("Bitrate (kbps)")
    plt.title("Bitrate Over Time")
    plt.legend()

    plt.tight_layout()
    plt.draw()

    chart_path = os.path.join(
        output_dir, "chart_" + ts_file_name + "_" + codec_type + stream_id + ".png"
    )
    plt.savefig(chart_path, bbox_inches="tight")

    # plt.show()
    plt.close()
    return chart_path


def process_streams(
    ts_file,
    audio_pids,
    video_pids,
    ts_file_name,
    output_dir,
    output_dict,
    audio_charts,
    video_charts,
    stream_info,
    media_info_av_data,
):
    """
    Process Audio and Video streams
    """
    audio_pts = process_audio_streams(
        ts_file,
        audio_pids,
        ts_file_name,
        output_dir,
        output_dict,
        audio_charts,
        stream_info,
        media_info_av_data,
    )
    video_pts = process_video_streams(
        ts_file,
        video_pids,
        ts_file_name,
        output_dir,
        output_dict,
        video_charts,
        media_info_av_data,
    )
    return audio_pts, video_pts


def validate_frame_duration(dict, frames):
    """
    Validates if frame duration is constant or variable
    """
    first_frame_duration = frames[0]["duration"]
    variable_duration = False

    for frame in frames:
        if first_frame_duration != frame["duration"]:
            variable_duration = True
            break

    if variable_duration == False:
        frame_duration = (first_frame_duration) / 90 / 1000
        fsec = int(frame_duration)
        fmillisec = int((frame_duration - fsec) * 1000)
        dict["Frame Duration"] = "{}s {}ms".format(fsec, fmillisec)
    else:
        dict["Frame Duration"] = "Variable Frame Duration Detected"


def validate_delay_bw_first_audio_and_video(dict, ts_file, stream_index):
    """
    Compute A/V interleaving
    """
    packets = get_all_packets(ts_file)

    previous_video_packet = None
    first_video_packet = None

    for packet in packets:
        if packet["codec_type"] == "video":
            first_video_packet = packet
            break

    for packet in packets:
        if packet["codec_type"] == "audio" and packet["stream_index"] == stream_index:
            break
        elif packet["codec_type"] == "video":
            previous_video_packet = packet

    first_av_delay = float(previous_video_packet["pts_time"]) - float(
        first_video_packet["pts_time"]
    )

    if first_av_delay > 0:
        dict["First audio packet arrival w.r.t video"] = (
            f"{int(first_av_delay)}s {int((first_av_delay - int(first_av_delay)) * 1000)}ms"
        )


def process_audio_streams(
    ts_file,
    audio_pids,
    ts_file_name,
    output_dir,
    output_dict,
    audio_charts,
    stream_info,
    media_info_av_data,
):
    """
    Process audio streams to perform AD analysis
    """
    audio_pts = []
    for i, pid in enumerate(audio_pids):
        frame_info = extract_frame_information(ts_file, pid)

        dump_json(
            frame_info,
            os.path.join(output_dir, f"audio_frames_info_{ts_file_name}_{pid}.json"),
        )

        output_dict["Ad_Analysis"]["Audio"].append({})
        validate_missing_packets(output_dict["Ad_Analysis"]["Audio"][i], pid, ts_file)
        output_dict["Ad_Analysis"]["Audio"][i]["ID"] = pid
        validate_frame_duration(output_dict["Ad_Analysis"]["Audio"][i], frame_info)
        audio_stream_idx = [
            stream.get("index") for stream in stream_info if pid == stream.get("id")
        ]
        validate_delay_bw_first_audio_and_video(
            output_dict["Ad_Analysis"]["Audio"][i], ts_file, audio_stream_idx[0]
        )
        audio_pts.append(
            get_asset_duration(
                "audio", pid, output_dict["Ad_Analysis"]["Audio"][i], ts_file
            )
        )
        bitrates = calculate_bitrate(frame_info)
        # Plot packet sizes and bitrates
        audio_charts.append(
            plot_pkt_sizes_and_bitrates(
                frame_info,
                bitrates,
                "audio",
                pid,
                output_dict["Ad_Analysis"]["Audio"][i],
                ts_file_name,
                output_dir,
            )
        )

    return audio_pts


def process_video_streams(
    ts_file,
    video_pids,
    ts_file_name,
    output_dir,
    output_dict,
    video_charts,
    media_info_av_data,
):
    """
    Process video streams to perform AD analysis
    """
    video_pts = []
    for j, pid in enumerate(video_pids):
        frame_info = extract_frame_information(ts_file, pid)
        dump_json(
            frame_info,
            os.path.join(output_dir, f"video_frames_info_{ts_file_name}_{pid}.json"),
        )

        output_dict["Ad_Analysis"]["Video"].append({})
        validate_missing_packets(output_dict["Ad_Analysis"]["Video"][j], pid, ts_file)
        output_dict["Ad_Analysis"]["Video"][j]["ID"] = pid
        validate_frame_duration(output_dict["Ad_Analysis"]["Video"][j], frame_info)
        video_pts.append(
            get_asset_duration(
                "video", pid, output_dict["Ad_Analysis"]["Video"][j], ts_file
            )
        )
        output_dict["Ad_Analysis"]["Video"][j]["GOP"] = frame_info[0]["pict_type"]

        bitrates = calculate_bitrate(frame_info)
        # Plot packet sizes and bitrates
        video_charts.append(
            plot_pkt_sizes_and_bitrates(
                frame_info,
                bitrates,
                "video",
                pid,
                output_dict["Ad_Analysis"]["Video"][j],
                ts_file_name,
                output_dir,
            )
        )

    return video_pts


def flatten_json(y):
    """
    Function to flatten JSON to store in excel
    """
    out = {}

    def flatten(x, name=""):
        if isinstance(x, dict):
            for a in x:
                flatten(x[a], name + a + "_")
        elif isinstance(x, list):
            out[name[:-1]] = ", ".join(map(str, x))
        else:
            out[name[:-1]] = x

    flatten(y)
    return out


def set_column_widths(sheet):
    """
    Function to set column widths in excel sheet
    """
    min_width, max_width, fixed_width = 15, 60, 40
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        sheet.column_dimensions[column].width = (
            adjusted_width if min_width <= max_length + 2 <= max_width else fixed_width
        )


def add_remarks_column_to_specific_sheets(
    file_path, sheet_name_prefix, remark_generator, media_info_av_data
):
    """
    Function to add remarks column to specific sheets
    """
    excel_data = pd.read_excel(file_path, sheet_name=None)
    for sheet_name, df in excel_data.items():
        if sheet_name.startswith(sheet_name_prefix):
            # df['Remarks'] = df.apply(remark_generator, axis=1)
            df["Remarks"] = df.apply(
                lambda row: remark_generator(row, df, sheet_name, media_info_av_data),
                axis=1,
            )

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        for sheet_name, df in excel_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def insert_charts_into_excel(
    output_path, video_charts, audio_charts, img_width=800, img_height=400
):
    """
    Function to insert charts into Excel
    """
    workbook = load_workbook(output_path)
    for i, chart_path in enumerate(video_charts):
        worksheet = workbook.create_sheet(f"Video_Chart_{i+1}")
        img = Image(chart_path)
        img.anchor, img.width, img.height = "B1", img_width, img_height
        worksheet.add_image(img)
    for i, chart_path in enumerate(audio_charts):
        worksheet = workbook.create_sheet(f"Audio_Chart_{i+1}")
        img = Image(chart_path)
        img.anchor, img.width, img.height = "B1", img_width, img_height
        worksheet.add_image(img)
    workbook.save(output_path)
    workbook.close()


def get_track_type(sheet_name):
    """
    Helper function to extract track type from excel sheet name
    """
    return "Audio" if "Audio" in sheet_name else "Video"


def get_track_num(sheet_name):
    """
    Helper function to extract track number from excel sheet name
    """
    return re.search(r"\d+", sheet_name).group()


def get_remarks(remark):
    """
    Helper function to return remarks
    """
    return "Ok" if remark == "Ok" else "Not Ok"


def filter_remarks(
    excel_path,
    audio_data,
    video_data,
    supported_audio_formats,
    supported_ac3_channels,
    expected_video_track_count,
    expected_audio_track_count,
):
    """
    Filters remarks based on the processed data and writes an overview sheet.
    """
    xls = pd.ExcelFile(excel_path)
    # filtered_rows = []

    av_tracks_as_expected = False

    # Helper function to create blank row
    def blank_row():
        return pd.DataFrame(
            [{"Type": "", "Value": "", "Remarks": "", "Sheet Name": ""}]
        )

    # A/V Track Count Section
    av_track_info = {
        "Video Tracks": len(video_data),
        "Audio Tracks": len(audio_data),
        "Audio Formats": [audio["Format"] for audio in audio_data],
        "Video Formats": [video["Format"] for video in video_data],
    }

    # All Sections
    av_section = []
    prop_section = []
    boundary_section = []
    gop_section = []
    bitrate_section = []
    frame_duration_section = []
    av_interleaving_section = []
    modified_audio_formats = []

    if len(video_data) != expected_video_track_count:
        av_section.append(
            {
                "Type": "Video Tracks ",
                "Value": str(len(video_data)),
                "Remarks": "Not Ok",
                "Sheet Name": "Overview",
            }
        )
        av_tracks_as_expected = False
    if len(audio_data) != expected_audio_track_count:
        av_section.append(
            {
                "Type": "Audio Tracks ",
                "Value": str(len(audio_data)),
                "Remarks": "Not Ok",
                "Sheet Name": "Overview",
            }
        )
    if len(audio_data) == expected_audio_track_count:
        for audio in audio_data:
            not_supported_audio = audio["Format"] not in supported_audio_formats
            not_supported_ac3 = (
                audio["Format"] == "AC-3"
                and audio["Channel(s)"] not in supported_ac3_channels
            )

            if (
                audio["Format"] == "AC-3"
                and audio["Channel(s)"] in supported_ac3_channels
            ):
                modified_audio_formats.append(
                    f"{audio['Format']} {audio['Channel(s)']}"
                )
            else:
                modified_audio_formats.append(audio["Format"])

            if not_supported_audio:
                av_section.append(
                    {
                        "Type": "Audio Tracks ",
                        "Value": audio["Format"] + " format not supported",
                        "Remarks": "Not Ok",
                        "Sheet Name": "Overview",
                    }
                )
                av_tracks_as_expected = False
                # break
            if not_supported_ac3:
                av_section.append(
                    {
                        "Type": "Audio Tracks ",
                        "Value": audio["Channel(s)"] + " not supported for AC-3",
                        "Remarks": "Not Ok",
                        "Sheet Name": "Overview",
                    }
                )
                av_tracks_as_expected = False
                # break
        # Update the dictionary
        av_track_info["Audio Formats"] = modified_audio_formats

    # if not av_section:
    av_section.append(
        {
            "Type": "A/V Tracks",
            "Value": str(av_track_info),
            "Remarks": "Ok",
            "Sheet Name": "Overview",
        }
    )

    if len(audio_data) == expected_audio_track_count and order_of_audio_tracks:
        current_order = [audio["Format"] for audio in audio_data]
        expected_order = sorted(
            current_order,
            key=lambda format: order_of_audio_tracks.get(format, float("inf")),
        )
        if current_order == expected_order:
            av_section.append(
                {
                    "Type": "Audio Tracks Order ",
                    "Value": str(current_order),
                    "Remarks": "Ok",
                    "Sheet Name": "Overview",
                }
            )
        else:
            av_section.append(
                {
                    "Type": "Audio Tracks Order ",
                    "Value": f"{str(current_order)} order is not matching expected {str(expected_order)}",
                    "Remarks": "Not Ok",
                    "Sheet Name": "Overview",
                }
            )

    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name)
        if "Remarks" in df.columns:
            for _, row in df.iterrows():
                track_type = get_track_type(sheet_name)
                track_num = get_track_num(sheet_name)
                if "Property" in row["Remarks"]:
                    value = (
                        row["Remarks"]
                        if pd.isna(row.get("Actual Value", 0))
                        else row.get("Actual Value", 0)
                    )
                    prop_section.append(
                        {
                            "Type": f"{track_type} Track {track_num} Properties",
                            "Value": value,
                            "Remarks": "Not Ok",
                            "Sheet Name": sheet_name,
                        }
                    )
                elif row.get("Key") in [
                    "Audio Boundary check at beginning",
                    "Audio Boundary check at the end",
                ]:
                    boundary_section.append(
                        {
                            "Type": f'{row["Key"]} For {track_type} Track {track_num}',
                            "Value": row["Value"],
                            "Remarks": get_remarks(row["Remarks"]),
                            "Sheet Name": sheet_name,
                        }
                    )
                elif row.get("Key") == "Missing Packets":
                    frame_duration_section.append(
                        {
                            "Type": f'{row["Key"]} For {track_type} Track {track_num}',
                            "Value": row["Value"],
                            "Remarks": get_remarks(row["Remarks"]),
                            "Sheet Name": sheet_name,
                        }
                    )
                elif row.get("Key") == "First audio packet arrival w.r.t video":
                    av_interleaving_section.append(
                        {
                            "Type": f'{row["Key"]} For {track_type} Track {track_num}',
                            "Value": row["Value"],
                            "Remarks": get_remarks(row["Remarks"]),
                            "Sheet Name": sheet_name,
                        }
                    )
                elif row.get("Key", 0) == "GOP":
                    gop_section.append(
                        {
                            "Type": "GOP Intact",
                            "Value": row["Value"],
                            "Remarks": "Ok" if row["Value"] == "I" else "Not Ok",
                            "Sheet Name": sheet_name,
                        }
                    )
                elif row.get("Key", 0) == "Average Bitrate":
                    kbps_value = float(row["Value"].split()[0])
                    mbps_value = (kbps_value * 1_000) / (8 * 1_024 * 1_024)
                    is_within_range = 2.8 <= mbps_value <= 3.0
                    bitrate_section.append(
                        {
                            "Type": f"Average Bitrate For {track_type} Track {track_num}",
                            "Value": row["Value"],
                            "Remarks": "Ok" if is_within_range == True else "Not Ok",
                            "Sheet Name": sheet_name,
                        }
                    )
                elif row.get("Key", 0) == "Frame Duration":
                    frame_duration_section.append(
                        {
                            "Type": f"Frame Duration for {track_type} Track {track_num}",
                            "Value": row["Value"],
                            "Remarks": (
                                "Ok" if "Variable" not in row["Value"] else "Not Ok"
                            ),
                            "Sheet Name": sheet_name,
                        }
                    )
                elif row.get("Key", 0) == "Max Frame size":
                    frame_duration_section.append(
                        {
                            "Type": f"Frame Size for {track_type} Track {track_num}",
                            "Value": row["Value"],
                            "Remarks": get_remarks(row["Remarks"]),
                            "Sheet Name": sheet_name,
                        }
                    )
                elif row.get("Key", 0) == "Duration":
                    boundary_section.append(
                        {
                            "Type": f"Total Duartion for {track_type} Track {track_num}",
                            "Value": row["Value"],
                            "Remarks": get_remarks(row["Remarks"]),
                            "Sheet Name": sheet_name,
                        }
                    )

    # Add sections for each type of verification and group the rows accordingly
    prop_section_df = pd.DataFrame(prop_section)
    prop_section_df.loc[-1] = [
        "1. A/V Property Analysis",
        "",
        "",
        "",
    ]  # Add section header
    prop_section_df.index = prop_section_df.index + 1
    prop_section_df = prop_section_df.sort_index()
    prop_section_df.drop_duplicates(
        subset=["Type", "Value", "Remarks", "Sheet Name"], inplace=True
    )

    av_section_df = pd.DataFrame(av_section)
    av_section_df.loc[-1] = [
        "2. AD Asset Component Verification",
        "",
        "",
        "",
    ]  # Add section header
    av_section_df.index = av_section_df.index + 1
    av_section_df = av_section_df.sort_index()
    av_section_df.drop_duplicates(
        subset=["Type", "Value", "Remarks", "Sheet Name"], inplace=True
    )

    boundary_section_df = frame_duration_section_df = gop_section_df = (
        av_interleaving_section_df
    ) = bitrate_section_df = pd.DataFrame()
    if len(boundary_section) > 0:
        boundary_section_df = pd.DataFrame(boundary_section)
        boundary_section_df.loc[-1] = [
            "3. A/V Boundary Verification",
            "",
            "",
            "",
        ]  # Add section header
        boundary_section_df.index = boundary_section_df.index + 1
        boundary_section_df = boundary_section_df.sort_index()
        boundary_section_df.drop_duplicates(
            subset=["Type", "Value", "Remarks", "Sheet Name"], inplace=True
        )

    if len(frame_duration_section) > 0:
        frame_duration_section_df = pd.DataFrame(frame_duration_section)
        frame_duration_section_df.loc[-1] = [
            "4. A/V Frame Verification",
            "",
            "",
            "",
        ]  # Add section header
        frame_duration_section_df.index = frame_duration_section_df.index + 1
        frame_duration_section_df = frame_duration_section_df.sort_index()
        frame_duration_section_df.drop_duplicates(
            subset=["Type", "Value", "Remarks", "Sheet Name"], inplace=True
        )

    if len(gop_section) > 0:
        gop_section_df = pd.DataFrame(gop_section)
        gop_section_df.loc[-1] = [
            "5. GOP Verification",
            "",
            "",
            "",
        ]  # Add section header
        gop_section_df.index = gop_section_df.index + 1
        gop_section_df = gop_section_df.sort_index()

    if len(av_interleaving_section) > 0:
        av_interleaving_section_df = pd.DataFrame(av_interleaving_section)
        av_interleaving_section_df.loc[-1] = [
            "6. A/V Interleaving Verification",
            "",
            "",
            "",
        ]  # Add section header
        av_interleaving_section_df.index = av_interleaving_section_df.index + 1
        av_interleaving_section_df = av_interleaving_section_df.sort_index()
        av_interleaving_section_df.drop_duplicates(
            subset=["Type", "Value", "Remarks", "Sheet Name"], inplace=True
        )

    if len(bitrate_section) > 0:
        bitrate_section_df = pd.DataFrame(bitrate_section)
        bitrate_section_df.loc[-1] = [
            "7. Bitrate Verification",
            "",
            "",
            "",
        ]  # Add section header
        bitrate_section_df.index = bitrate_section_df.index + 1
        bitrate_section_df = bitrate_section_df.sort_index()
        bitrate_section_df.drop_duplicates(
            subset=["Type", "Value", "Remarks", "Sheet Name"], inplace=True
        )

    # Combine all sections with blank rows in between
    combined_df = pd.concat(
        [
            prop_section_df,
            blank_row(),
            av_section_df,
            blank_row(),
            boundary_section_df,
            blank_row(),
            frame_duration_section_df,
            blank_row(),
            gop_section_df,
            blank_row(),
            av_interleaving_section_df,
            blank_row(),
            bitrate_section_df,
        ],
        ignore_index=True,
    )

    with pd.ExcelWriter(
        excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        combined_df.to_excel(writer, sheet_name="Overview", index=False)

    book = load_workbook(excel_path)
    sheets = book.sheetnames
    book._sheets = [book["Overview"]] + [
        book[sheet] for sheet in sheets if sheet != "Overview"
    ]
    book.save(excel_path)


def convert_to_milliseconds(time_str):
    """
    Helper function to convert time in string format '1s 20ms" format to ms
    """
    seconds_match = re.search(r"(\d+)\s*s", time_str)
    milliseconds_match = re.search(r"(\d+)\s*ms", time_str)

    total_ms = 0
    if seconds_match:
        total_ms += int(seconds_match.group(1)) * 1000
    if milliseconds_match:
        total_ms += int(milliseconds_match.group(1))
    return total_ms


def is_duration_exceeding_threshold(extra_duration_str, frame_duration_str):
    """
    Validate duration is exceeding threshold
    """
    extra_duration_total_ms = convert_to_milliseconds(extra_duration_str)
    frame_duration_total_ms = convert_to_milliseconds(frame_duration_str)

    return abs(extra_duration_total_ms) > frame_duration_total_ms


def remark_generator(row, df, sheet_name, media_info_av_data):
    """
    Define a remark generator function based on specific conditions
    """
    if (
        row.get("Key") in ["Missing Packets", "Missing Packets Count"]
        and str(row.get("Value")) != "0"
    ):
        return "Missing packets detected"
    if row.get("Key") in [
        "Audio Boundary check at beginning",
        "Audio Boundary check at the end",
    ]:
        duration_str = row["Value"]
        # Validate if frame duration exceeds threshold
        exceeds_threshold = False
        frame_duration_row = df[df["Key"] == "Frame Duration"]
        if not frame_duration_row.empty:
            frame_duration_value = frame_duration_row.iloc[0]["Value"]
            exceeds_threshold = is_duration_exceeding_threshold(
                duration_str, frame_duration_value
            )
        direction = "ahead of" if "beginning" in row.get("Key") else "lagging behind"
        return (
            f'Audio is {direction} video by {row["Value"]}'
            if exceeds_threshold
            else "Ok"
        )
    if row.get("Key") == "GOP" and row.get("Value") != "I":
        return "First frame is not I-Frame"
    if row.get("Key") == "First audio packet arrival w.r.t video":
        return (
            "There is " + row["Value"] + " delay between first video and audio packet"
            if convert_to_milliseconds(row["Value"]) > 50
            else "Ok"
        )
        # return 'There is ' + row['Value'] +'s delay between first video and audio packet' if float(row['Value'])*1000> 50.0 else 'Ok'
    if row.get("Key") == "Frame Duration":
        return "Ok" if "Variable" not in row["Value"] else "Not Ok"
    if row.get("Key", 0) == "Max Frame size":
        max_frame_size = float(re.search(r"\d+(\.\d+)?", row.get("Value")).group())
        if "Audio" in sheet_name:
            return "Frame size exceeds threshold" if max_frame_size > 200.0 else "Ok"
        else:
            return "Frame size exceeds threshold" if max_frame_size > 1024.0 else "Ok"
    if row.get("Key", 0) == "Duration":
        remark = "Not Ok"
        total_ad_duration = media_info_av_data["General"][0]["Duration"]
        total_ad_dur_in_ms_frm_mediaInfo = convert_to_milliseconds(total_ad_duration)
        track_duration_in_ms = convert_to_milliseconds(row.get("Value"))

        trackType = get_track_type(sheet_name)
        trackNum = get_track_num(sheet_name)

        track_dur_frm_mediaInfo = (
            media_info_av_data[trackType][int(trackNum) - 1].get("Duration", 0)
            if int(trackNum) - 1 < len(media_info_av_data[trackType])
            else "0s 0ms"
        )
        track_dur_frm_mediaInfo_in_ms = convert_to_milliseconds(track_dur_frm_mediaInfo)

        if (
            track_duration_in_ms <= total_ad_dur_in_ms_frm_mediaInfo
            and track_dur_frm_mediaInfo_in_ms == track_duration_in_ms
        ):
            remark = "Ok"
        if track_duration_in_ms > total_ad_dur_in_ms_frm_mediaInfo:
            remark = f"{trackType} {trackNum} duration is more than total AD Duration of {total_ad_duration}"
        if track_dur_frm_mediaInfo_in_ms != track_duration_in_ms:
            remark = f"{trackType} {trackNum} duration is not equal to MediaInfo duration of {track_dur_frm_mediaInfo}"
        return remark
    return "Ok"


def check_audio_boundaries(audio_pts, video_pts, output_dict):
    """
    Compute boundary values for Audio w.r.t video
    """
    for video_item in video_pts:
        for i, audio_item in enumerate(audio_pts):
            first_pts_diff = (
                (video_item.get("First Frame PTS") - audio_item.get("First Frame PTS"))
                / 90
                / 1000
            )
            last_pts_diff = (
                (audio_item.get("Last Frame PTS") - video_item.get("Last Frame PTS"))
                / 90
                / 1000
            )

            output_dict["Ad_Analysis"]["Audio"][i][
                "Audio Boundary check at beginning"
            ] = f"{int(first_pts_diff)}s {int((first_pts_diff - int(first_pts_diff)) * 1000)}ms"
            output_dict["Ad_Analysis"]["Audio"][i][
                "Audio Boundary check at the end"
            ] = f"{int(last_pts_diff)}s {abs(int((last_pts_diff - int(last_pts_diff)) * 1000))}ms"


def get_asset_duration(codec_type, stream_id, dict, ts_file):
    """
    Extract Duration for Audio and Video Streams
    """
    frames = get_pts_duration_from_packet_info(ts_file, stream_id)
    sorted_frames = sorted(frames, key=lambda x: int(x.get("pts", "0")))

    first_frame_pts_time = sorted_frames[0]["pts_time"]
    last_frame_pts_time = sorted_frames[-1]["pts_time"]

    duration = (
        float(last_frame_pts_time)
        - float(first_frame_pts_time)
        + float(sorted_frames[-1]["duration_time"])
    )
    dict["First Frame PTS"] = str(sorted_frames[0]["pts"])
    dict["Last Frame PTS"] = str(sorted_frames[-1]["pts"])
    dict["Duration"] = f"{int(duration)}s {int((duration - int(duration)) * 1000)}ms"

    return {
        "codec_type": codec_type,
        "pid": stream_id,
        "First Frame PTS": sorted_frames[0]["pts"],
        "Last Frame PTS": sorted_frames[-1]["pts"],
    }


def validate_missing_packets(dict, pid, ts_file):
    """
    Identify Missing Packets
    """
    packets = get_pts_duration_from_packet_info(ts_file, pid)
    missing_packets = []
    prev_pts = None
    prev_duration = None

    sorted_packets = sorted(packets, key=lambda x: int(x.get("pts", "0")))

    for packet in sorted_packets:
        pts = packet.get("pts", 0)
        duration = packet.get("duration", 0)

        if prev_pts is not None:
            # Check for missing packets between current and previous packet
            if pts > prev_pts + prev_duration:
                missing_start = prev_pts + prev_duration
                missing_end = pts
                missing_packets.extend(range(missing_start, missing_end, prev_duration))

        # Update previous packet information
        prev_pts = pts
        prev_duration = duration

    dict["Missing Packets"] = missing_packets if missing_packets else [0]
    dict["Missing Packets Count"] = (
        str(len(missing_packets)) if missing_packets else "0"
    )


def calculate_bitrate(frames):
    """
    Calculates bitrate over time
    """
    if not frames:
        return []

    bitrates = []
    window_size = 1  # 1 second window
    start_time = float(frames[0].get("pts_time", 0))
    end_time = float(frames[-1].get("pts_time", 0))

    current_time = start_time
    while current_time < end_time:
        window_frames = [
            frame
            for frame in frames
            if current_time
            <= float(frame.get("pts_time", 0))
            < current_time + window_size
        ]
        total_size = sum(int(frame.get("pkt_size", 0)) for frame in window_frames)
        bitrate = (total_size * 8) / window_size
        bitrates.append((current_time, bitrate))
        current_time += window_size

    return bitrates


if __name__ == "__main__":
    main()
