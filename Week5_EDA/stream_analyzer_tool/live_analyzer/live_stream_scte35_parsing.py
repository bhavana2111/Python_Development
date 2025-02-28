import struct, re
import sys, json, os
import subprocess, shutil
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

# Get the parent directory of the current file
parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
# Add the parent directory to sys.path
sys.path.insert(0, parent_dir)
import config.project_config as config
from utils.media_info import extract_media_info
from utils.frame_info import extract_pts_frame_type

TS_PACKET_SIZE = 188
TS_PACKET_SIZE_HEX_CHARS = 376
TS_HEADER_SIZE = 4
SPLICE_TABLE_ID = 252
MC_SWTSP_MAX_PID = 8192
PES_PTS_DTS_INDICATOR_MASK = 192
TS_PES_PTS_OFFSET = 18 # 9*2
TS_PES_DTS_OFFSET = 28 # 14*2

SWTSP_TS_ADAPDATION_FIELD_MASK = 3
SWTSP_TS_ADAPDATION_FIELD_ONLY = 2
SWTSP_TS_ADAPDATION_FIELD_PAYLOAD = 3
SWTSP_TS_ADAPDATION_FIELD_MAX_LEN = 183
SWTSP_TS_PCR_FLAG_MASK = 16
SWTSP_TS_RAI_FLAG_MASK = 0x40 # Random Access Indicator flag

ADAPTATION_FIELD_FLAG_BIT = 0x30
ADAPTATION_FIELD_PLUS_PAYLOAD = 0x30
ADAPTATION_FIELD_NO_PAYLOAD = 0x20
CHECK_PTS_FLAG_BIT = 0x80
CHECK_DTS_FLAG_BIT = 0x40

ADAPT_LEN_FIELD_OFFSET = 4  # Offset of adaptation length field from start of TS packet
ADAPT_LEN_FIELD_LENGTH = 1  # Length of field specifying adaptation field length
PES_PTS_FIELD_START_OFFSET = 9  # Offset of PTS field from start of PES
PES_DTS_FIELD_START_OFFSET = 14  # Offset of DTS field from start of PES
PES_PTS_DTS_FLAGS_FIELD_OFFSET = 7  # Offset of byte carrying PTS and DTS avail flag from start of PES

PTS_CLOCK_FREQUENCY = 90000  # 90 kHz

#PAT - Global Variables
is_pat_found            = False
pat_section_dump        = ''
pat_service_list            = []
pmt_pid_list_in_pat         = []

#PMT Global Variables
pmt_section_data_dump       = []
list_of_pmt_pids            = []
pmt_sections                = []
scd_service_details_in_pmt  = []
sync_byte_recovery_pos      = []
sync_byte_error_info        = []
PMT_TABLE_ID                = "0x02"

#Common Variables
SCTE_TABLE_ID           = "0xFC"
scte_35_pid_and_count   = []
scte_35_sections        = []
scte_35_pid_data        = []
scte_35_sections_with_pid   = []
scte_35_pid_and_pgm_num = []
scte_35_video_pid_mapping = {}
video_pid_isScrambled   = {}
scte_35_endPts          = {}
scte_35_onids           = []
scte_35_dur_error_info  = []
scte_35_clash_info      = []
unique_scte35_ids       = set()

tot_splice_insert_msg       = 0
tot_splice_null_msg         = 0
tot_splice_bw_reserve       = 0
tot_splice_time_signal      = 0
tot_unknown_splice_type   = 0

#debug variables:
enable_PAT_debugs       = config.enable_debugs["PAT"]
enable_PMT_debugs       = config.enable_debugs["PMT"]
enable_SCTE_debugs      = config.enable_debugs["SCTE"]


pid_occurrences = {}
pid_data = {}

# Global variables
project_name    = config.project_name
loop_packet_info = []
wrapAround_packet_info = []
noSplicehit_packet_info = {}
base_utc_time = 0
splice_service_infos = []
stream_timeline_data = {
        "stream_start_utc" : "",
        "stream_end_utc" : "",
        "splice_hit_utc" : [],
        "spot_utc" : []
    }

def load_standard_values(file_path):
    file_path = os.path.join(parent_dir, 'config', file_path)
    with open(file_path, 'r') as file:
        return json.load(file)

# Load project specific configs
standard_values = load_standard_values(config.project_spec_json[project_name])
keys_not_expected = standard_values.get('keys_not_expected', {})
expected_audio_track_count = standard_values.get('expected_audio_track_count', 0)
expected_video_track_count = standard_values.get('expected_video_track_count', 0)
supported_audio_formats = standard_values.get('supported_audio_formats', [])
supported_ac3_channels = standard_values.get('supported_ac3_channels', [])
order_of_audio_tracks = standard_values.get('order_of_audio_tracks',{})

# Function to calculate adaptation and ts header size
def get_adapt_ts_header(buf):
    adapt = (buf[3] >> 4) & 0x03
    if adapt == 0:
        ts_header_size = TS_PACKET_SIZE
    elif adapt == 1:
        ts_header_size = TS_HEADER_SIZE
    elif adapt == 2:
        ts_header_size = TS_PACKET_SIZE
    elif adapt == 3:
        ts_header_size = TS_HEADER_SIZE + buf[4] + 1
    else:
        ts_header_size = TS_PACKET_SIZE
    return ts_header_size

def calculate_pts_dts(current_packet):
    pts = 0
    is_rai = 0
    if(
        (current_packet[3] & ADAPTATION_FIELD_FLAG_BIT) == ADAPTATION_FIELD_PLUS_PAYLOAD
        or (current_packet[3] & ADAPTATION_FIELD_FLAG_BIT) == ADAPTATION_FIELD_NO_PAYLOAD
    ):
        adaptation_length = current_packet[ADAPT_LEN_FIELD_OFFSET]
        is_rai = (current_packet[ADAPT_LEN_FIELD_OFFSET + 1] & SWTSP_TS_RAI_FLAG_MASK) >> 6
        pts_start_offset = (
            TS_HEADER_SIZE + ADAPT_LEN_FIELD_LENGTH + PES_PTS_FIELD_START_OFFSET + adaptation_length
        )
        dts_start_offset = (
            TS_HEADER_SIZE + ADAPT_LEN_FIELD_LENGTH + PES_DTS_FIELD_START_OFFSET + adaptation_length
        )
        pts_dts_avail_offset = (
            TS_HEADER_SIZE + ADAPT_LEN_FIELD_LENGTH + PES_PTS_DTS_FLAGS_FIELD_OFFSET + adaptation_length
        )
    else:
        pts_start_offset = TS_HEADER_SIZE + PES_PTS_FIELD_START_OFFSET
        dts_start_offset = TS_HEADER_SIZE + PES_DTS_FIELD_START_OFFSET
        pts_dts_avail_offset = TS_HEADER_SIZE + PES_PTS_DTS_FLAGS_FIELD_OFFSET

    if current_packet[pts_dts_avail_offset] & CHECK_PTS_FLAG_BIT:
        pts = (
            ((current_packet[pts_start_offset] & 0x0E) << 29)
            | (current_packet[pts_start_offset + 1] << 22)
            | ((current_packet[pts_start_offset + 2] & 0xFE) << 14)
            | (current_packet[pts_start_offset + 3] << 7)
            | (current_packet[pts_start_offset + 4] >> 1)
        )

    if current_packet[pts_dts_avail_offset] & CHECK_DTS_FLAG_BIT:
        dts = (
            ((current_packet[dts_start_offset] & 0x0E) << 29)
            | (current_packet[dts_start_offset + 1] << 22)
            | ((current_packet[dts_start_offset + 2] & 0xFE) << 14)
            | (current_packet[dts_start_offset + 3] << 7)
            | (current_packet[dts_start_offset + 4] >> 1)
        )

        pts = dts

    return pts, is_rai

# Convert PTS to seconds
def convert_pts_to_sec(pts):
    return int(pts / PTS_CLOCK_FREQUENCY)

# Convert PTS to UTC time
def convert_pts_to_utc(pts):
    global base_utc_time
    utc_time = 0

    # Convert PTS to seconds
    #time_in_seconds = round(pts / PTS_CLOCK_FREQUENCY, 2)
    time_in_seconds = pts / PTS_CLOCK_FREQUENCY
    # Calculate & return UTC time
    if base_utc_time != 0:
        utc_time = base_utc_time + timedelta(seconds=time_in_seconds)
        #utc_time = utc_time.replace(microsecond=0)
        #utc_time = utc_time.replace(microsecond=int(utc_time.microsecond / 1000) * 1000)
        utc_time = utc_time.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

    return utc_time

def parse_pat_section():
    global pat_section_dump
    global pat_service_list

    loc_sec_dump_copy = pat_section_dump

    table_id = loc_sec_dump_copy[0]
    loc_sec_dump_copy = loc_sec_dump_copy[1:]

    section_length = ((loc_sec_dump_copy[0] & 0x0F) << 8) | loc_sec_dump_copy[1]
    loc_sec_dump_copy = loc_sec_dump_copy[2:]

    transport_stream_id = ((loc_sec_dump_copy[0] & 0xFF) << 8) | loc_sec_dump_copy[1]
    loc_sec_dump_copy   = loc_sec_dump_copy[2:]

    version_number = (loc_sec_dump_copy[0] & 0x3E)>>1
    loc_sec_dump_copy = loc_sec_dump_copy[1:]

    section_number      = loc_sec_dump_copy[0]
    last_section_number = loc_sec_dump_copy[1]
    loc_sec_dump_copy = loc_sec_dump_copy[2:]

    if enable_PAT_debugs == True:
        print("Table ID            : [0x%x] "%(table_id))
        print("Section Length      : [%d] "%(section_length))
        print("Transport Stream ID : [0x%x] "%(transport_stream_id))
        print("Version Number      : [%d] "%(version_number))
        print("Section Number      : [%d] "%(section_number))
        print("Last Section Number : [%d] "%(last_section_number))

    total_length = section_length - 9 #5=Table Header and 4=CRC

    length = 0
    while length < total_length:
        program_number  = ((loc_sec_dump_copy[0] & 0xFF) << 8) | loc_sec_dump_copy[1]
        program_map_PID = ((loc_sec_dump_copy[2] & 0x1F) << 8) | loc_sec_dump_copy[3]
        pat_service_list.append([int(program_number),int(program_map_PID)])
        pmt_pid_list_in_pat.append(program_map_PID)
        loc_sec_dump_copy = loc_sec_dump_copy[4:]
        if enable_PAT_debugs == True:
            print("    SID :[%d] PID :[%d]" %(program_number,program_map_PID))
        length = length + 4

def parse_pmt_section():

    global pmt_sections
    global scte_35_pid_and_pgm_num
    global scte_35_video_pid_mapping
    video_pid = 0

    if enable_PMT_debugs == True:
        print("Total PMT Sections " +str(len(pmt_sections)))

    for index in range(len(pmt_sections)):

        loc_sec_dump_copy   = pmt_sections[index]

        table_id            = loc_sec_dump_copy[0]
        loc_sec_dump_copy   = loc_sec_dump_copy[1:]

        section_length      = ((loc_sec_dump_copy[0] & 0x0F) << 8) | loc_sec_dump_copy[1]
        loc_sec_dump_copy   = loc_sec_dump_copy[2:]

        program_number      = ((loc_sec_dump_copy[0] & 0xFF) << 8) | loc_sec_dump_copy[1]
        loc_sec_dump_copy   = loc_sec_dump_copy[2:]

        version_number      = (loc_sec_dump_copy[0] & 0x3E)>>1
        loc_sec_dump_copy   = loc_sec_dump_copy[1:]

        section_number      = loc_sec_dump_copy[0]
        last_section_number = loc_sec_dump_copy[1]
        loc_sec_dump_copy   = loc_sec_dump_copy[2:]

        pcr_pid             = ((loc_sec_dump_copy[0] & 0x1F) << 8) | loc_sec_dump_copy[1]
        loc_sec_dump_copy   = loc_sec_dump_copy[2:]

        program_info_length = ((loc_sec_dump_copy[0] & 0x0F) << 8) | loc_sec_dump_copy[1]
        loc_sec_dump_copy   = loc_sec_dump_copy[2:]

        if program_info_length > 0:
            loc_sec_dump_copy = loc_sec_dump_copy[(program_info_length):]

        if enable_PMT_debugs == True:
            print("Table ID            : [0x%x] "%(table_id))
            print("Section Length      : [%d] "%(section_length))
            print("Program Number      : [%d] [0x%x] "%(program_number,program_number))
            print("Version Number      : [%d] "%(version_number))
            print("Section Number      : [%d] "%(section_number))
            print("Last Section Number : [%d] "%(last_section_number))
            print("PCR PID             : [%d] "%(pcr_pid))
            print("Program Info Length : [%d] "%(program_info_length))

        total_length = section_length - 9 - program_info_length - 4 #9=Table Header and 4=CRC

        length = 0

        while length < total_length:
            stream_type         = loc_sec_dump_copy[0]
            loc_sec_dump_copy   = loc_sec_dump_copy[1:]

            elementary_PID      = ((loc_sec_dump_copy[0] & 0x1F) << 8) | loc_sec_dump_copy[1]
            loc_sec_dump_copy   = loc_sec_dump_copy[2:]

            es_info_length      = int(((loc_sec_dump_copy[0] & 0x0F) << 8) | loc_sec_dump_copy[1])
            loc_sec_dump_copy   = loc_sec_dump_copy[2:]

            if enable_PMT_debugs == True:
                print("    Stream Type    : [%d]"%(stream_type))
                print("    Elementary PID : [%d]"%(elementary_PID))
                print("    ES Info Length : [%d]"%(es_info_length))

            # Storing video pid using stream type 27 (0x18)
            if stream_type == 27:
                video_pid = elementary_PID

            scte_35_pid_present = False
            # SCTE_35 type is 134 (0x86)
            if stream_type == 134:
                #if enable_SCTE_debugs == True:
                #    print("SCTE_35 PID[%d] Identified in PGM[%d]"%(elementary_PID,program_number))
                scte_35_pid_present = True

            inner_desc_loop_length = 0
            while inner_desc_loop_length < es_info_length:
                inner_descriptor_tag    = loc_sec_dump_copy[0]
                inner_descriptor_length = int(loc_sec_dump_copy[1])

                #CueIdentifier descriptor checking
                if scte_35_pid_present == True and inner_descriptor_tag == 138:
                    #if enable_SCTE_debugs == True:
                    #    print("Adding PID[%d] in SCTE-35 List"%(elementary_PID))
                    pid_and_pgm_num = str(program_number) +","+str(elementary_PID)
                    scte_35_pid_and_pgm_num.append(pid_and_pgm_num)
                    scte_35_video_pid_mapping[elementary_PID] = video_pid
                    scte_35_pid_present = False

                if enable_PMT_debugs == True:
                    print("        Desc Tag           : [%d] [0x%x]"%(inner_descriptor_tag,inner_descriptor_tag))
                    print("        Desc Length        : [%d]"%(inner_descriptor_length))
                inner_desc_loop_length = inner_desc_loop_length + int(inner_descriptor_length) + 2 #2 for desc_tag and desc_len
                loc_sec_dump_copy = loc_sec_dump_copy[(inner_descriptor_length+2):]
            length = length + int(es_info_length) + 5 #5 for stream_type(1),elementary_PID(2),ES_info_length(2)

def parse_scte35_sections():
    global scte_35_sections_with_pid
    global scte_35_pid_data
    global tot_splice_insert_msg
    global tot_splice_null_msg
    global tot_splice_bw_reserve
    global tot_splice_time_signal
    global tot_unknown_splice_type

    #print(f"scte_35_sections_with_pid = {scte_35_sections_with_pid}")
    for index in range(len(scte_35_sections_with_pid)):
        splice_pid = scte_35_sections_with_pid[index][0]
        if enable_SCTE_debugs == True:
            print("SCTE-35 PID[%d] Data"%(splice_pid))
            print("-------------------------")
        loc_sec_dump_copy   =  scte_35_sections_with_pid[index][2]
        table_id            = loc_sec_dump_copy[0]
        loc_sec_dump_copy   = loc_sec_dump_copy[1:]

        section_length      = ((loc_sec_dump_copy[0] & 0x0F) << 8) | loc_sec_dump_copy[1]
        loc_sec_dump_copy   = loc_sec_dump_copy[2:]

        protocol_version    = loc_sec_dump_copy[0]
        loc_sec_dump_copy   = loc_sec_dump_copy[1:]

        encrypted_packet        = ((loc_sec_dump_copy[0] & 0x80) >> 7)
        encryption_algorithm    = ((loc_sec_dump_copy[0] & 0x7E) >> 1)

        pts_adjustment          = ((loc_sec_dump_copy[0] & 0x01) << 32)  | ((loc_sec_dump_copy[1]) << 24) | ((loc_sec_dump_copy[2]) << 16) | ((loc_sec_dump_copy[3]) << 8) | (loc_sec_dump_copy[4])
        loc_sec_dump_copy       = loc_sec_dump_copy[5:]

        cw_index                = loc_sec_dump_copy[0]
        loc_sec_dump_copy       = loc_sec_dump_copy[1:]

        #reserved=0xFFF (or) Tier - 12 bits, currently dont have spec, so unable to confirm this field
        #Splice Command Length | 12 bits

        splice_command_length = ((loc_sec_dump_copy[1] & 0x0F) << 8)  | (loc_sec_dump_copy[2])
        loc_sec_dump_copy       = loc_sec_dump_copy[3:]

        splice_command_type     = (loc_sec_dump_copy[0])
        loc_sec_dump_copy       = loc_sec_dump_copy[1:]
        splice_command_type_str = ''
        if splice_command_type == 4 or splice_command_type == 0:
            splice_command_type_str = "Splice Null"
            tot_splice_null_msg     += 1
        elif splice_command_type == 5:
            splice_command_type_str = "Splice Insert"
            tot_splice_insert_msg   += 1
        elif splice_command_type == 6:
            splice_command_type_str = "Time Signal"
            tot_splice_time_signal   += 1
        elif splice_command_type == 7:
            splice_command_type_str = "Bandwidth Reservation"
            tot_splice_bw_reserve   += 1
        else:
            splice_command_type_str = "unknown, yet to check"
            tot_unknown_splice_type += 1

        if enable_SCTE_debugs == True:
            print("Table ID                 : 0x%x "%(table_id))
            print("Section Length           : %d "%(section_length))
            print("Protocol Version         : 0x%x "%(protocol_version))
            print("Encrypted Packet         : %d "%(encrypted_packet))
            print("Encryption Algorithm     : %d "%(encryption_algorithm))
            print("PTS Adjustment           : %d "%(pts_adjustment))
            print("CW Index                 : %d "%(cw_index))
            print("Splice Command Length    : %d "%(splice_command_length))
            print("Splice Command Type      : [%d] [%s]"%(splice_command_type,splice_command_type_str))

        if splice_command_type == 5:
            splice_event_id                 = ((loc_sec_dump_copy[0]) << 24) | ((loc_sec_dump_copy[1]) << 16) | ((loc_sec_dump_copy[2]) << 8) | (loc_sec_dump_copy[3])
            loc_sec_dump_copy               = loc_sec_dump_copy[4:]
            splice_event_cancel_indicator   = ((loc_sec_dump_copy[0] & 0x80) >> 7)
            loc_sec_dump_copy               = loc_sec_dump_copy[1:]

            if splice_event_cancel_indicator == 0:
                out_of_network_indicator        = ((loc_sec_dump_copy[0] & 0x80) >> 7)
                program_splice_flag             = ((loc_sec_dump_copy[0] & 0x40) >> 6)
                duration_flag                   = ((loc_sec_dump_copy[0] & 0x20) >> 5)
                splice_immediate_flag           = ((loc_sec_dump_copy[0] & 0x10) >> 4)
                #4bits reserved
                loc_sec_dump_copy               = loc_sec_dump_copy[1:]

                if splice_immediate_flag == 0:
                    #Parse spliceTime - 5bytes
                    time_specified_flag             = ((loc_sec_dump_copy[0] & 0x80) >> 7)
                    #next 6-bits reserved
                    #pts_time = 33 bits
                    pts_time                        = ((loc_sec_dump_copy[0] & 0x01) << 32)  | ((loc_sec_dump_copy[1]) << 24) | ((loc_sec_dump_copy[2]) << 16) | ((loc_sec_dump_copy[3]) << 8) | (loc_sec_dump_copy[4])
                    loc_sec_dump_copy               = loc_sec_dump_copy[5:]

                    if duration_flag == 1:

                        #Parse breakduration - 5bytes
                        auto_return                     = ((loc_sec_dump_copy[0] & 0x80) >> 7)
                        #next 6-bits reserved
                        #duration = 33 bits
                        duration                        = ((loc_sec_dump_copy[0] & 0x01) << 32)  | ((loc_sec_dump_copy[1]) << 24) | ((loc_sec_dump_copy[2]) << 16) | ((loc_sec_dump_copy[3]) << 8) | (loc_sec_dump_copy[4])
                        duration_in_secs                = (duration/90)/1000
                        loc_sec_dump_copy               = loc_sec_dump_copy[5:]

                    #unique_program_id = 16bits
                    unique_program_id               = ((loc_sec_dump_copy[0] & 0xFF) << 8)  | (loc_sec_dump_copy[1])
                    loc_sec_dump_copy               = loc_sec_dump_copy[2:]

                    #avail_num = 8bits
                    avail_num                       = loc_sec_dump_copy[0]
                    loc_sec_dump_copy               = loc_sec_dump_copy[1:]

                    #avails_expected = 8bits
                    avails_expected                 = loc_sec_dump_copy[0]
                    loc_sec_dump_copy               = loc_sec_dump_copy[1:]

                    descriptor_loop_length          = ((loc_sec_dump_copy[0] & 0xFF) << 8)  | (loc_sec_dump_copy[1])
                    loc_sec_dump_copy               = loc_sec_dump_copy[2:]

                    if descriptor_loop_length > 0:
                        descriptor_tag          = loc_sec_dump_copy[0]
                        loc_sec_dump_copy       = loc_sec_dump_copy[1:]
                        descriptor_length       = loc_sec_dump_copy[0]
                        loc_sec_dump_copy       = loc_sec_dump_copy[1:]
                        identifier              = ((loc_sec_dump_copy[0]) << 24) | ((loc_sec_dump_copy[1]) << 16) | ((loc_sec_dump_copy[2]) << 8) | (loc_sec_dump_copy[3])
                        identifier_str          = chr(loc_sec_dump_copy[0]) + chr(loc_sec_dump_copy[1]) + chr(loc_sec_dump_copy[2]) + chr(loc_sec_dump_copy[3])
                        loc_sec_dump_copy       = loc_sec_dump_copy[4:]
                        provider_avail_id       = ((loc_sec_dump_copy[0]) << 24) | ((loc_sec_dump_copy[1]) << 16) | ((loc_sec_dump_copy[2]) << 8) | (loc_sec_dump_copy[3])
                        loc_sec_dump_copy       = loc_sec_dump_copy[4:]

            if enable_SCTE_debugs == True:
                print("Splice Event Id                  : [0x%x][%d] "%(splice_event_id,splice_event_id))
                print("Splice Event Cancel Indicator    : %d "%(splice_event_cancel_indicator))
                if splice_event_cancel_indicator == 0:
                    print("Out of Network Indicator         : %d "%(out_of_network_indicator))
                    print("Program Splice Flag              : %d "%(program_splice_flag))
                    print("Duration Flag                    : %d "%(duration_flag))
                    print("Splice Immediate Flag            : %d "%(splice_immediate_flag))
                    if splice_immediate_flag == 0:
                        print("Time Specified Flag              : %d "%(time_specified_flag))
                        print("PTS Time                         : %d "%(pts_time))
                        if duration_flag == 1:
                            print("Auto Return                      : %d "%(auto_return))
                            print("Duration                         : %d [%d secs] "%(duration,duration_in_secs))
                        print("Unique Program Id                : %d "%(unique_program_id))
                        print("Avail Num                        : %d "%(avail_num))
                        print("Avails Expected                  : %d "%(avails_expected))
                        print("Desc Loop Length                 : %d "%(descriptor_loop_length))

                        if descriptor_loop_length > 0:
                            print("Descriptor Tag                  : [0x%x][%d] "%(descriptor_tag,descriptor_tag))
                            print("Descriptor Length               : %d "%(descriptor_length))
                            print("Identifier                      : [0x%x] [%s]"%(identifier,identifier_str))
                            print("Provider Avail Id               : %d "%(provider_avail_id))

            temp_splice_data = {
                "spot_count"        : len(scte_35_pid_data[str(splice_pid)]) + 1,
                "splice_eventId"    : splice_event_id,
                "splice_pid"        : splice_pid,
                "video_pid"         : scte_35_sections_with_pid[index][1],
                "splice_cancel"     : splice_event_cancel_indicator,
                "splice_immediate"  : splice_immediate_flag
            }

            if splice_event_cancel_indicator == 0:
                temp_splice_data["first_pes_pts"]   = scte_35_sections_with_pid[index][3]
                temp_splice_data["splice_hit_pts"]  = scte_35_sections_with_pid[index][4]
                temp_splice_data["onid"]            = out_of_network_indicator
                if splice_immediate_flag == 0:
                    temp_splice_data["start_pts"]       = pts_time
                    if duration_flag == 1:
                        temp_splice_data["duration"] = duration
            scte_35_pid_data[str(splice_pid)].append(temp_splice_data)
        if enable_SCTE_debugs == True:
            print("\n")

def parse_scte_35_packet(p_packet_data):
    global scte_35_sections
    global scte_35_sections_with_pid
    global loop_packet_info

    pid = struct.unpack('>H', p_packet_data[1:3])[0] & 0x1FFF
    pusi = (int(p_packet_data[1])& 0x40)>>6
    adaptation_field_control = (int(p_packet_data[3]) & 0x30)>>4
    payload_offset = 4

    if adaptation_field_control & 0x1 != 0:
        if pusi == 1:
            table_id = int(p_packet_data[5])
            section_length  = ((p_packet_data[6] & 0x0F) << 8) | p_packet_data[7]
            payload_offset  = 5

            # Adding 3 to consider table_id and section_length
            last_index      = (section_length + 3 )

            #Moving to payload start position
            p_packet_data   = p_packet_data[payload_offset:]

            #table_id '0xFC'= 252
            if table_id == 252:
                tmp_packet_data = p_packet_data[:last_index-4] #without CRC
                if tmp_packet_data not in scte_35_sections:
                    scte_35_sections.append(tmp_packet_data)
                    scte_35_data            = tmp_packet_data
                    scte_35_pid_and_data    = [pid, scte_35_video_pid_mapping[pid], scte_35_data]
                    scte_35_sections_with_pid.append(scte_35_pid_and_data)
                    if pid in pid_data:
                        pid_data[pid]["data"].append(tmp_packet_data)
                    else:
                        pid_data[pid] = {"data": [tmp_packet_data]}
                else:
                    # Storing error loop packet data available from the stream for final report
                    loop_packet_data = [pid, scte_35_video_pid_mapping[pid]]
                    if loop_packet_data not in loop_packet_info:
                        loop_packet_info.append(loop_packet_data)

def is_section_already_added(table_id):
    global pmt_section_data_dump

    section_found       = False
    local_section_list  = []

    if table_id == PMT_TABLE_ID:
        section_number_in_new_section   = int(pmt_section_data_dump[6])
        program_number                  = ((pmt_section_data_dump[3] & 0xFF) << 8) | pmt_section_data_dump[4]
        local_section_list              = pmt_sections

    for index in range(len(local_section_list)):
        local_section                   = local_section_list[index]
        section_num_in_exist_list       = int(local_section[6])
        pgm_num_in_exist_list           = ((local_section[3] & 0xFF) << 8) | local_section[4]
        if section_num_in_exist_list == section_number_in_new_section and program_number == pgm_num_in_exist_list:
            section_found   = True
            break
    return section_found

def is_pmt_pid_exists_in_list(pid):
    global list_of_pmt_pids

    pid_exist_status    = False
    total_pmt_pids      = len(list_of_pmt_pids)

    if total_pmt_pids > 0:
        for index in range(len(list_of_pmt_pids)):
            pid_value   = list_of_pmt_pids[index]
            if int(pid_value) == int(pid):
                pid_exist_status    = True
                break
    return pid_exist_status

def is_pes_header_scrambled(ts_packet):
    # Assuming ts_packet is a byte array representing a single 188-byte TS packet
    scrambling_control = (ts_packet[3] & 0xC0) >> 6
    if scrambling_control == 0:
        return "PES Header (Clear)"
    elif scrambling_control == 2 or scrambling_control == 3:
        return "PES Header (Scrambled)"
    else:
        return "PES Header (Reserved)"

def prepare_pmt_table():
    global pmt_sections
    global pmt_section_data_dump
    is_sec_exists   = is_section_already_added(PMT_TABLE_ID)
    if is_sec_exists == False:
        section_number_in_new_section   = int(pmt_section_data_dump[6])
        program_number_in_new_section   = ((pmt_section_data_dump[3] & 0xFF) << 8) | pmt_section_data_dump[4]
        if enable_PMT_debugs == True:
            print("Adding PMT PGM[%d] SectionNo[%d] "%(program_number_in_new_section,section_number_in_new_section))
        pmt_sections.append(pmt_section_data_dump)

def prepare_pmt_section(p_packet_data):
    global pmt_section_data_dump
    global pmt_total_sec_length
    global is_pmt_first_packet_received
    global list_of_pmt_pids

    pid             = struct.unpack('>H', p_packet_data[1:3])[0] & 0x1FFF
    packet_size     = TS_PACKET_SIZE
    payload_offset  = 4
    pusi            = (int(p_packet_data[1]) & 0x40) >> 6
    adap_fld_ctrl   = (int(p_packet_data[3]) & 0x30) >> 4

    if adap_fld_ctrl & 0x1 != 0:
        if pusi == 1:
            pmt_section_data_dump = []
            payload_offset  = 5
            table_id        = int(p_packet_data[5])
            section_length  = ((p_packet_data[6] & 0x0F) << 8) | p_packet_data[7]
            section_number  = p_packet_data[11]
            pmt_total_sec_length            = section_length + 3 #3 Added for table_id and section_length field
            is_pmt_first_packet_received    = True

        if pmt_total_sec_length > packet_size:
            end_position            = packet_size
            #pmt_total_sec_length    = pmt_total_sec_length - (packet_size)
            pmt_total_sec_length   = pmt_total_sec_length - (188 - payload_offset)
        else:
            #Need to remove the hardcoded value 4 bytes
            end_position           = pmt_total_sec_length + 4
            #end_position            = pmt_total_sec_length
            pmt_total_sec_length    = 0

        if is_pmt_first_packet_received == True:

            pmt_section_data_dump.extend(p_packet_data[payload_offset:])

            if pmt_total_sec_length == 0 and len(pmt_section_data_dump) != 0:
                prepare_pmt_table()
                list_of_pmt_pids.append(pid)
                pmt_section_data_dump           = []
                is_pmt_first_packet_received    = False

def parse_pat_packet(p_packet_data):
    global pat_section_dump
    global glob_tsid_in_pat

    global is_pmt_first_packet_received

    pid = struct.unpack('>H', p_packet_data[1:3])[0] & 0x1FFF

    if pid == 0 and enable_PAT_debugs == True:
        print("Adding PAT Section ")

    pusi                        = (int(p_packet_data[1])& 0x40)>>6
    adaptation_field_control    = (int(p_packet_data[3]) & 0x30)>>4
    payload_offset              = 4

    if adaptation_field_control & 0x1 != 0:
        if pusi == 1:
            table_id                        = int(p_packet_data[5])
            section_length                  = ((p_packet_data[6] & 0x0F) << 8) | p_packet_data[7]
            section_number                  = p_packet_data[11]
            payload_offset                  = 5
            is_pmt_first_packet_received    = True

            # Adding 3 to consider table_id and section_length
            last_index      = (section_length + 3 )

            #Moving to payload start position
            p_packet_data   = p_packet_data[payload_offset:]

            glob_tsid_in_pat    = ((p_packet_data[3] & 0xFF) << 8) | p_packet_data[4]
            pat_section_dump    = p_packet_data[:last_index]

def find_next_sync_byte(ts_file):
    while True:
        byte = ts_file.read(1)
        if not byte:
            return None
        if hex(byte[0]) == '0x47' or hex(byte[0]) == '0xB8':
            return ts_file.tell() - 1  # Return position of sync byte

def seek_to_next_proper_sync_byte(ts_file):
    # Find the position of the next sync byte
    new_position = find_next_sync_byte(ts_file)
    if new_position is None:
        return False, None

    ts_file.seek(new_position)
    #print(f"Found sync byte at position: {new_position}")

    # Check next 8 packets to ensure correct sync
    valid_stream = True
    for _ in range(8):
        packet_data = ts_file.read(TS_PACKET_SIZE)
        if not packet_data or (hex(packet_data[0]) != '0x47' and hex(packet_data[0]) != '0xB8'):
            valid_stream = False
            #print("Sync byte not found in validation packets, searching again...")
            ts_file.seek(new_position + 1)  # Continue searching after the initial found sync byte
            break

    return valid_stream, new_position

def parse_stream_to_prepare_pat_and_pmt(file_name):
    global is_pat_found
    global pmt_pid_list_in_pat
    global sync_byte_recovery_pos
    global sync_byte_error_info

    packet_size = TS_PACKET_SIZE  # Standard packet size in bytes

    with open(file_name, 'rb') as ts_file:
        while True:
            packet_data = ts_file.read(packet_size)
            if not packet_data:
                break

            sync_byte = hex(packet_data[0])
            if sync_byte != '0x47' and sync_byte != '0xB8':
                current_position = ts_file.tell()
                #print("[PMT] Sync byte not found, searching for next sync byte...")
                # Go back by the number of bytes read minus one, to recheck the current position
                ts_file.seek(-len(packet_data) + 1, 1)

                is_valid_stream, new_position = seek_to_next_proper_sync_byte(ts_file)
                if new_position is None:
                    sync_byte_error_info.append("No more sync bytes found, ending process.")
                    break
                elif not is_valid_stream:
                    continue
                else:
                    sync_byte_recovery_pos.append(new_position)
                    sync_byte_error_info.append(f"Sync byte error found at [{current_position}] and moved to new position [{new_position}]")
                    #print("[PMT] Valid sync sequence found, continuing processing...")
                    #print("\n")

            # Parse the PID from the transport stream header
            pid = struct.unpack('>H', packet_data[1:3])[0] & 0x1FFF

            if pid == 0 and is_pat_found == False:  # PAT PID = 0
                parse_pat_packet(packet_data)
                parse_pat_section()
                is_pat_found = True

            if is_pat_found == True:
                pmt_pid_exists_in_list  = is_pmt_pid_exists_in_list(pid)
                if pid in pmt_pid_list_in_pat and pmt_pid_exists_in_list == False:
                    if project_name != "osn" or pid != 16:
                        prepare_pmt_section(packet_data)

def parse_stream_for_scte35_msg(file_name, scte_35_pid_list, scte_35_video_pid_list):
    global pid_occurrences
    global scte_35_sections_with_pid
    global wrapAround_packet_info
    global noSplicehit_packet_info
    global video_pid_isScrambled
    global scte_35_video_pid_mapping
    global sync_byte_recovery_pos
    first_iframe_pes_pts = {}
    iframe_pes_pts = 0
    scte_pid_recvd = 0
    local_splice_pid = 0
    pkt_cnt = 0
    noSplicehit_packet_info_added = False

    if len(scte_35_pid_list) == 0:
        print("[ERROR] SCTE-35 PIDs are not Present in this stream")
        sys.exit(1)

    packet_size = TS_PACKET_SIZE  # Standard packet size in bytes
    total_scte_35_packets = 0

    with open(file_name, 'rb') as ts_file:
        while True:
            current_packet = ts_file.read(packet_size)
            if not current_packet:
                break

            # Parse the sync_byte from the transport stream header
            sync_byte = hex(current_packet[0])
            if sync_byte != '0x47' and sync_byte != '0xB8':
                # Go back by the number of bytes read minus one, to recheck the current position
                ts_file.seek(-len(current_packet) + 1, 1)

                if len(sync_byte_recovery_pos) > 0:
                    if ts_file.tell() < sync_byte_recovery_pos[0]:
                        # Seek to next valid sync byte packet position
                        ts_file.seek(sync_byte_recovery_pos[0])
                        sync_byte_recovery_pos = sync_byte_recovery_pos[1:]
                        current_packet = ts_file.read(packet_size)
                else:
                    is_valid_stream, new_position = seek_to_next_proper_sync_byte(ts_file)
                    if new_position is None:
                        break
                    elif not is_valid_stream:
                        continue
                    else:
                        print("[SCTE] Valid sync sequence found, continuing processing...")

            # Parse the PID from the transport stream header
            pid = struct.unpack('>H', current_packet[1:3])[0] & 0x1FFF
            pusi = (int(current_packet[1]) & 0x40) >> 6
            ts_header_size = get_adapt_ts_header(current_packet)

            if pid in scte_35_pid_list:
                # When next splice point data received before previous splice point data is processed
                if scte_pid_recvd == 1:
                    if pid not in noSplicehit_packet_info:
                        noSplicehit_packet_info[pid] = []
                    noSplicehit_packet_info[pid].append(total_scte_35_packets)
                    noSplicehit_packet_info_added = True

                total_scte_35_packets += 1
                scte_pid_recvd = 1
                local_splice_pid = pid
                if pid in pid_occurrences:
                    pid_occurrences[pid] += 1  # Increment the occurrence count
                else:
                    pid_occurrences[pid] = 1  # Initialize the occurrence count
                parse_scte_35_packet(current_packet)

            # If there is payload
            elif pid < MC_SWTSP_MAX_PID and pid in scte_35_video_pid_list:
                if pid not in video_pid_isScrambled:
                    video_pid_isScrambled[pid] = is_pes_header_scrambled(current_packet)

                if ( ts_header_size < TS_PACKET_SIZE and len(current_packet) <= TS_PACKET_SIZE
                    and current_packet[ts_header_size] == 0
                    and current_packet[ts_header_size + 1] == 0
                    and current_packet[ts_header_size + 2] == 1
                    and pusi == 1
                ):
                    pes_pts, is_rai = calculate_pts_dts(current_packet)

                    iframe_pes_pts = 0
                    if is_rai == 1:
                        if pid not in first_iframe_pes_pts:
                            first_iframe_pes_pts[pid] = pes_pts
                        elif pes_pts < first_iframe_pes_pts[pid]:
                            first_iframe_pes_pts[pid] = pes_pts
                            wa_packet_data = [local_splice_pid, pid]
                            if wa_packet_data not in wrapAround_packet_info:
                                wrapAround_packet_info.append(wa_packet_data)
                        iframe_pes_pts = pes_pts

                    if scte_pid_recvd == 1 and pid == scte_35_video_pid_mapping[local_splice_pid]:
                        scte_pid_recvd = 0
                        splice_hit_pts = iframe_pes_pts if iframe_pes_pts != 0 else pes_pts
                        if total_scte_35_packets > 0:
                            scte_35_sections_with_pid[total_scte_35_packets - 1].extend([first_iframe_pes_pts[pid],splice_hit_pts])
                        if total_scte_35_packets > 1 and noSplicehit_packet_info_added:
                            # Since unable to find splice_hit_pts for previous spot, assign current spot hit pts to previous aswell
                            noSplicehit_packet_info_added = False
                            scte_35_sections_with_pid[total_scte_35_packets - 2].extend([first_iframe_pes_pts[pid],splice_hit_pts])

def parse_scte_summary_str(input_str):
    # Extract everything between square brackets using regex
    matches = re.findall(r'\[(.*?)\]', input_str)
    # Assign extracted values to separate variables
    parsed_data = {
        "spot_no"       :   int(input_str.split()[0]),
        "onid"          :   int(matches[0]),  # "1"
        "spot_start_utc":   matches[1],  # "UTC 2024-08-02 08:50:40"
        "spot_start_sec":   matches[2],  # "62.6 sec"
        "spot_pts"      :   int(matches[3]),  # "3807190580"
        "pict_type"     :   matches[4],  # "I"
        "splice_hit_utc":   matches[5],  # "UTC 2024-08-02 08:50:34"
        "break_time"    :   matches[6]  # "6.32 sec"
    }

    return parsed_data

def display_summary_data_in_pair(data, key, sections):
    global scte_35_dur_error_info
    global stream_timeline_data

    # Helper function to parse the time from the string
    def parse_time(data_string):
        # Find the position of the '[UTC ' string, which precedes the date and time
        start_index = data_string.find('[UTC ') + len('[UTC ')
        end_index = data_string.find(']', start_index)

        # Extract the date and time portion
        time_str = data_string[start_index:end_index]

        # Parse the extracted date and time
        return datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S.%f')

    def format_summary_data(data):
        formatted_summary = []

        # Determine the number of columns by checking the first line
        max_columns = len(data[0].split('['))

        # Initialize max_lengths list to accommodate all columns
        max_lengths = [0] * max_columns

        # Find the maximum lengths of each column
        for line in data:
            columns = line.split('[')
            for i, col in enumerate(columns):
                max_lengths[i] = max(max_lengths[i], len(col.strip()))

        # Display the aligned output
        for line in data:
            columns = line.split('[')
            formatted_line = ''
            for i, col in enumerate(columns):
                formatted_line += col.strip().ljust(max_lengths[i] + 5) + '['
            # Remove the last unnecessary '['
            formatted_line = formatted_line.rstrip('[')
            formatted_summary.append(formatted_line)
        return formatted_summary

    data = format_summary_data(data)

    # Array to store spurious splice data and its duration
    spurious_splice_data = []

    # Array to store the matched onid[1] and onid[0]
    matched_data = []
    inner_loop_break = False
    # Traverse each ONID [1] in the data
    i = 0
    while i < len(data):
        if "[1]" in data[i]:
            #if len(data) == 1:
            #    data.pop(i)
            #    break
            time1 = parse_time(data[i])
            matched_data.append(data[i])
            j = i + 1

            while j < len(data):
                if "[0]" in data[j]:
                    time0 = parse_time(data[j])
                    time_diff = int((time0 - time1).total_seconds())
                    # Condition: time difference should be one of [10,15,20,30,40,60] which is in config file
                    if time_diff in config.allowed_ad_duration:
                        matched_data.append(data[j])
                        # Remove the matched ONID [1] and ONID [0] from data
                        data.pop(j)
                        data.pop(i)
                        inner_loop_break = True
                        break
                    else:
                        # Remove unsatisfied ONID [0] and push to spurious_splice_data
                        spurious_splice_data.append(data.pop(j))
                        scte_35_dur_error_info.append(time_diff)
                        j -= 1  # Adjust index after removal

                j += 1
        else:
            # Remove unsatisfied ONID [0] and push to spurious_splice_data
            spurious_splice_data.append(data.pop(i))
            i = 0
            continue
        if inner_loop_break:
            i = 0
            continue
        i += 1

    # Any remaining ONID [0] that doesn't match conditions
    i = 0
    while i < len(data):
        if "[0]" in data[i]:
            spurious_splice_data.append(data.pop(i))
        else:
            i += 1

    # Processing & Displaying all perfect splice data
    for i in range(len(matched_data)):
        dur_str = ""
        set_msg = False
        b2b_spot = False
        manual_dur_set = False
        curr_summary = {}
        next_summary = {}
        next_spot_cnt = -1

        j = i + 1

        curr_summary = parse_scte_summary_str(matched_data[i])
        stream_timeline_data["splice_hit_utc"].append(curr_summary['splice_hit_utc'])
        stream_timeline_data["spot_utc"].append(curr_summary['spot_start_utc'])
        spot_cnt = curr_summary['spot_no'] - 1
        spot_section = sections[spot_cnt]

        if "[1]" in matched_data[i]:
            time1 = parse_time(matched_data[i])
            dur_in_secs = int(spot_section['duration'] / PTS_CLOCK_FREQUENCY) if "duration" in spot_section else 30 # manually setting to 30sec if duration pair is not available
            manual_dur_set = True
        else:
            dur_in_secs = 0 # manually setting to 0sec if ONID is "0"
            manual_dur_set = False
        if j < len(matched_data):
            next_summary = parse_scte_summary_str(matched_data[j])
            next_spot_cnt = next_summary['spot_no'] - 1
            if "[0]" in matched_data[j]:
                time0 = parse_time(matched_data[j])
                dur_in_secs = int((time0 - time1).total_seconds())
                manual_dur_set = False

        # Calculating duration and end_pts for all spot sections
        if spot_section['onid'] == curr_summary['onid'] and spot_section['start_pts'] == curr_summary['spot_pts']:
            spot_section['duration'] = f"{dur_in_secs} sec"
        if spot_section['end_pts'] is None:
            spot_section['end_pts'] = curr_summary['spot_pts'] + (dur_in_secs * PTS_CLOCK_FREQUENCY)

        # Validating whether there is any clash packet data
        if next_spot_cnt > 0 and sections[next_spot_cnt]["start_pts"] < spot_section["end_pts"]:
            err_str = f"    {next_spot_cnt}. Splice PID [{key}] has clash spot data with current start PTS [{sections[next_spot_cnt]['start_pts']}] < previous end PTS [{spot_section['end_pts']}]"
            scte_35_clash_info.append(err_str)

        if i % 2 == 0 or manual_dur_set:
            set_msg = True
            if i > 1:
                print("\n")
        else:
            set_msg = False

        # calculate LIVE duration till the first splice data
        beginning_live_dur = int((time1 - base_utc_time).total_seconds())
        if i == 0 and beginning_live_dur > 0:
            dur_str += f"               /* Playing LIVE for {beginning_live_dur} seconds */\n\n"

        # Calculate LIVE duration for rest of the spot gap data and B2B spot information
        if (i % 2 == 0 or manual_dur_set) and i > 0:
            prev_time0 = parse_time(matched_data[i-1])
            gap_time = int((time1 - prev_time0).total_seconds())
            gap_time -= dur_in_secs if manual_dur_set else 0
            if gap_time > 0 and set_msg:
                dur_str += f"               /* Playing LIVE for {gap_time} seconds */\n\n"
            else:
                b2b_spot = True

        if set_msg:
            dur_str += f"               /* {dur_in_secs} sec Ad duration */"
            dur_str += f"  /* B2B Ad Spot */" if b2b_spot else ""
            spot_section['spot_type'] = "B2B" if b2b_spot else "Single"
            print(dur_str)
        print(matched_data[i])

    if len(spurious_splice_data) > 0:
        # Displaying all spurious splice data
        print("\nSpurious Splice Data:")
        print("-----------------------")
        for line in spurious_splice_data:
            print(line)

def prepare_scte35_summary_data(ts_file):
    global scte_35_pid_data
    global scte_35_onids
    global scte_35_endPts
    global loop_packet_info
    global wrapAround_packet_info
    global noSplicehit_packet_info
    global scte_35_dur_error_info
    global scte_35_clash_info
    global scte_35_video_pid_list
    global sync_byte_error_info
    pts_to_pict_type = {}

    for video_pid in scte_35_video_pid_list:
        scte_35_endPts[video_pid] = []

        # Get ffprobe output
        ffprobe_output = extract_pts_frame_type(ts_file, hex(video_pid))
        # Create a dictionary for quick lookup of pict_type by PTS
        pts_to_pict_type[video_pid] = {frame['pts']: frame['pict_type'] for frame in ffprobe_output}

    #print(f"scte_35_pid_data = {json.dumps(scte_35_pid_data)}")
    print("===================================================================")
    print("################ Summarized Output Values ############### ")
    print("===================================================================")
    for key, sections in scte_35_pid_data.items():
        onid_data = []
        scte35_summary = []
        print(f"\nSplice PID [{key}]")
        print("******************")
        if len(sections) > 0:
            print("-----------------------------------------------------------------------------------------------------------------------------------")
            print("Spot# ONID         Spot Start Time         FromStreamStart   SpotPTS        GOP        Splice hit time           Break Time(sec)")
            print("-----------------------------------------------------------------------------------------------------------------------------------")
            for idx, section in enumerate(sections):
                if ("splice_cancel" in section or "splice_immediate" in section):
                    if section["splice_cancel"]:
                        print(f"{idx+1}. Splice event cancel indicator is set")
                        scte_35_endPts[section["video_pid"]].append(0)
                        continue
                    if section["splice_immediate"]:
                        print(f"{idx+1}. Splice event immediate flag is set")
                        scte_35_endPts[section["video_pid"]].append(0)
                        continue

                onid_data.append(section["onid"])

                if idx > 0:
                    section["spot2spot_diff"] = f"{round(section["start_pts"] / PTS_CLOCK_FREQUENCY,2) - round(sections[idx-1]["start_pts"] / PTS_CLOCK_FREQUENCY,2)} sec"
                else:
                    section["spot2spot_diff"] = "0.0 sec"
                spot_pts_pict_type = pts_to_pict_type[scte_35_video_pid_mapping[int(key)]].get(section["start_pts"], None)
                spot_from_video_start_pts = section["start_pts"] - section["first_pes_pts"]
                spot_from_video_start_utc = convert_pts_to_utc(spot_from_video_start_pts)
                spot_from_video_start_sec = convert_pts_to_sec(spot_from_video_start_pts)

                splice_hit_from_video_start_pts =  section["splice_hit_pts"] - section["first_pes_pts"]
                splice_hit_time_utc = convert_pts_to_utc(splice_hit_from_video_start_pts)

                spot_to_splice_hit_pts = section["start_pts"] - section["splice_hit_pts"]
                spot_to_splice_hit_sec = convert_pts_to_sec(spot_to_splice_hit_pts)

                summary_str = f"{idx+1} [{section['onid']}] [UTC {spot_from_video_start_utc}] [{spot_from_video_start_sec}s] [{section['start_pts']}] [{spot_pts_pict_type}] [UTC {splice_hit_time_utc}] [{spot_to_splice_hit_sec}s]"
                scte35_summary.append(summary_str)

                if "duration" not in section:
                    section["end_pts"] = None # end_pts is calculated below based on ONID [1] & [0]
                else:
                    section["end_pts"] = section["start_pts"] + section["duration"]

            display_summary_data_in_pair(scte35_summary, key, sections)
        else:
            print("     No Splice Data available")
        scte_35_onids.append({key : onid_data})

    print("\n")
    print("===================================================================")
    print(" SCTE-35 Error Report ")
    print("===================================================================\n")

    # Display the error checklist report
    print("Error Checklist")
    print("---------------")
    print(f"Sync Byte check --------------------------- [ {'FAIL' if len(sync_byte_error_info) > 0 else 'OK'} ]")
    print(f"Duration check ---------------------------- [ {'FAIL' if len(scte_35_dur_error_info) > 0 else 'OK'} ]")
    print(f"Loop packet check ------------------------- [ {'FAIL' if len(loop_packet_info) > 0 else 'OK'} ]")
    print(f"Wrap around packet check ------------------ [ {'FAIL' if len(wrapAround_packet_info) > 0 else 'OK'} ]")
    print(f"Splice PES packet check ------------------- [ {'FAIL' if len(noSplicehit_packet_info) > 0 else 'OK'} ]")
    print(f"Splice clash check ------------------------ [ {'FAIL' if len(scte_35_clash_info) > 0 else 'OK'} ]")

    # Display sync byte error report
    if len(sync_byte_error_info) > 0:
        print("\nSync byte Error")
        print("---------------")
        for error_data in sync_byte_error_info:
            print(f"    {error_data}")

    # Display duration error report when not matching to 30secs if present in the stream
    if len(scte_35_dur_error_info) > 0:
        print("\nDuration Error")
        print("----------------")
        print(f"    Expected configured splice duration (in secs)   = {config.allowed_ad_duration}")
        print(f"    Received spurious splice duration (in secs)     = {scte_35_dur_error_info}")

    # Display loop packet error report if present in the stream
    if len(loop_packet_info) > 0:
        print("\nLoop Packet Error")
        print("-----------------")
        for idx,error_data in enumerate(loop_packet_info):
            print("     Splice PID [%d] Video PID [%d] Loop packets found."%(loop_packet_info[idx][0], loop_packet_info[idx][1]))

    # Display wrap around packet error report if present in the stream
    if len(wrapAround_packet_info) > 0:
        print("\nWrap around packet Error")
        print("------------------------")
        for idx,error_data in enumerate(wrapAround_packet_info):
            print("     Splice PID [%d] Video PID [%d] Wrap around packets found."%(wrapAround_packet_info[idx][0], wrapAround_packet_info[idx][1]))

    # Display splice hit message error for particular splice in the stream
    if len(noSplicehit_packet_info) > 0:
        print("\nSplice Message Time Error")
        print("-------------------------")
        for key, values in noSplicehit_packet_info.items():
            for val in values:
                print(f"    Splice PID [{key}] {val}th Splice packet has NO PUSI set before {val+1}th Splice data")

    # Display ONID error report if not represented as 1 followed 0 if present in the stream
    if len(scte_35_onids) > 0:
        def check_ones_followed_by_zeros():
            errors = []
            for dct in scte_35_onids:
                for key, value in dct.items():
                    if len(value) == 0:
                        errors.append(f"    Splice PID [{key}]: ONID's are not processed because of Empty/Splice Cancel/Immediate event")
                        continue
                    # Check if all elements are 1's or all elements are 0's
                    if all(v == 1 for v in value):
                        errors.append(f"    Splice PID [{key}]: Every ONID's are 1")
                        continue
                    elif all(v == 0 for v in value):
                        errors.append(f"    Splice PID [{key}]: Every ONID's are 0")
                        continue

                    for i in range(len(value) - 1):
                        if value[i] == 1 and value[i + 1] != 0:
                            errors.append(f"    {i+1}. Splice PID [{key}]: Every ONID's 1 must be followed by a 0")
                    if value[-1] == 1:
                        errors.append(f"    {i+1}. Splice PID [{key}]: Every ONID's 1 must be followed by a 0")
            if errors:
                return "\n".join(errors)
            else:
                return None

        error_str = check_ones_followed_by_zeros()
        if error_str is not None:
            print("\nONID 1 & 0 Error")
            print("----------------")
            print(error_str)

    # Display Spot clash error report if present in the stream
    if len(scte_35_clash_info) > 0:
        print("\nClash Spot data Error")
        print("---------------------")
        for error_data in scte_35_clash_info:
            print(error_data)

def transform_list_to_dict(input_list):
    return {str(item): [] for item in input_list}

# Function to extract the SERVICES ANALYSIS REPORT
def extract_tsp_analysis_report(output):
    lines = output.splitlines()

    transport_stream_report = []
    service_analysis_report = []

    in_transport_stream_report = False
    in_service_analysis_report = False

    for line in lines:
        if "TRANSPORT STREAM ANALYSIS REPORT" in line:
            in_transport_stream_report = True
            in_service_analysis_report = False
        elif "SERVICES ANALYSIS REPORT" in line:
            in_service_analysis_report = True
            in_transport_stream_report = False
        elif "===============================================================================" in line:
            in_transport_stream_report = False
            in_service_analysis_report = False

        if in_transport_stream_report:
            transport_stream_report.append(line)
        elif in_service_analysis_report:
            service_analysis_report.append(line)

    return '\n'.join(transport_stream_report), '\n'.join(service_analysis_report)

def process_wraparound_pkt(pts_list):
    # Logic to convert wrap around packet PTS to incremental PTS
    def handle_pts_wraparound(pts, previous_pts, wraparound_counter):
        max_pts = 2**33
        if pts < previous_pts and (previous_pts - pts) > (max_pts / 2):
            # Detected wrap-around
            wraparound_counter += 1
        full_pts = pts + wraparound_counter * max_pts
        return full_pts, wraparound_counter

    revised_pts = []
    wraparound_counter = 0
    previous_pts = 0
    for pts in pts_list:
        full_pts, wraparound_counter = handle_pts_wraparound(pts, previous_pts, wraparound_counter)
        revised_pts.append(full_pts)
        previous_pts = pts

    return revised_pts

def get_base_utc_time_from_stream(ts_report):
    global stream_timeline_data
    start_utc_time = 0
    # Regular expression to match the "First TDT UTC time stamp"
    stream_start_utc_match = re.search(r"First (?:TDT|STT) UTC time stamp:\s*\.*\s*([0-9]{4}/[0-9]{2}/[0-9]{2} [0-9]{2}:[0-9]{2}:[0-9]{2})", ts_report)

    if stream_start_utc_match:
        first_tdt_utc_str = stream_start_utc_match.group(1)
        # Parse the timestamp string to a datetime object
        first_tdt_utc = datetime.strptime(first_tdt_utc_str, "%Y/%m/%d %H:%M:%S")
        stream_timeline_data["stream_start_utc"] = f"UTC {first_tdt_utc}.{first_tdt_utc.microsecond}"
        # Extract the individual components and convert to a list
        start_utc_time = datetime(first_tdt_utc.year, first_tdt_utc.month, first_tdt_utc.day, first_tdt_utc.hour, first_tdt_utc.minute, first_tdt_utc.second)
    else:
        print("First TDT UTC time stamp not found.")

    # Regular expression to match the "Last TDT UTC time stamp"
    stream_end_utc_match = re.search(r"Last (?:TDT|STT) UTC time stamp:\s*\.*\s*([0-9]{4}/[0-9]{2}/[0-9]{2} [0-9]{2}:[0-9]{2}:[0-9]{2})", ts_report)

    if stream_end_utc_match:
        last_tdt_utc_str = stream_end_utc_match.group(1)
        # Parse the timestamp string to a datetime object
        last_tdt_utc = datetime.strptime(last_tdt_utc_str, "%Y/%m/%d %H:%M:%S")
        stream_timeline_data["stream_end_utc"] = f"UTC {last_tdt_utc}.{last_tdt_utc.microsecond}"
    else:
        print("Last TDT UTC time stamp not found.")

    return start_utc_time

def format_hex_pid(input_str):
    # Remove the '0x' and leading zeros from the input
    hex_value = input_str.lstrip('0x').upper()

    # Convert the hex value to its decimal equivalent
    decimal_value = int(hex_value, 16)

    # Format the output
    return f'{decimal_value} (0x{hex_value})'

def process_service_table(table):
    splice_service_info = {
        "video_pids": {
            "clear": [],
            "scrambled": [],
            "shared": []
        },
        "audio_pids": {
            "clear": [],
            "scrambled": [],
            "shared": []
        }
    }

    # Regular expressions to capture video and audio PIDs with their access type
    video_pid_regex = re.compile(r"(\b0x[0-9A-Fa-f]{3,4})\s+AVC video")
    audio_pid_regex = re.compile(r"(\b0x[0-9A-Fa-f]{3,4})\s+(AC-3 Audio|MPEG-2 Audio)")

    # Access type regex
    access_type_regex = re.compile(r"\s+([C|S|\+])\s")

    # Split the table into lines
    lines = table.strip().split('\n')

    # Iterate through the lines to find PIDs and their access types
    for line in lines:
        # Find video PIDs
        video_match = video_pid_regex.search(line)
        audio_match = audio_pid_regex.search(line)
        access_match = access_type_regex.search(line)

        if video_match and access_match:
            pid = video_match.group(1)
            access_type = access_match.group(1)
            if access_type == "C":
                splice_service_info['video_pids']['clear'].append(format_hex_pid(pid))
            elif access_type == "S":
                splice_service_info['video_pids']['scrambled'].append(format_hex_pid(pid))
            elif access_type == "+":
                splice_service_info['video_pids']['shared'].append(format_hex_pid(pid))

        # Find audio PIDs
        if audio_match and access_match:
            pid = audio_match.group(1)
            access_type = access_match.group(1)
            if access_type == "C":
                splice_service_info['audio_pids']['clear'].append(format_hex_pid(pid))
            elif access_type == "S":
                splice_service_info['audio_pids']['scrambled'].append(format_hex_pid(pid))
            elif access_type == "+":
                splice_service_info['audio_pids']['shared'].append(format_hex_pid(pid))

    return splice_service_info

def extract_service_tables_from_file(content):
    # Split the content based on table separators
    tables = re.split(r'\|=+\|', content)

    results = []

    for table in tables:
        # Check if the table contains "SCTE 35 Splice Info"
        if "SCTE 35 Splice Info" in table:
            # Process the table and store the results
            splice_service_info = process_service_table(table)
            results.append(splice_service_info)

    return results

def flatten_json(y):
    """
    Function to flatten JSON to store in excel
    """
    out = {}

    def flatten(x, name=''):
        if isinstance(x, dict):
            for a in x:
                flatten(x[a], name + a + '_')
        elif isinstance(x, list):
            out[name[:-1]] = ', '.join(map(str, x))
        else:
            out[name[:-1]] = x

    flatten(y)
    return out

def generate_remarks(data, standard_values,key_prefix, keys_not_expected=None):
    """
    Generates remarks for data based on standard values and keys not expected.
    """
    return [
        'Ok' if k in standard_values and str(v) == str(standard_values.get(k))
        else 'Ok' if k not in standard_values and v is not None and (not keys_not_expected or k not in keys_not_expected[key_prefix])
        else 'Property is not present' if v is None
        else 'Property not expected for this format' if keys_not_expected and k in keys_not_expected[key_prefix]
        else 'Property value is not as per standard. Expected is {}'.format(str(standard_values.get(k)))
        for k, v in data.items()
    ]

def generate_std_values(data, standard_values):
    """
    Generates standard values for data based on standard values from project config.
    """
    return [
        standard_values.get(k, str(data.get(k, 0)))
        for k in data.keys()
    ]

def process_media_data(media_data, standard_values, keys_not_expected=None, supported_formats=None, supported_channels=None):
    """
    Processes media data (video or audio) and writes it to a DataFrame.
    """
    media_json = []
    media_df = []
    remarks_column_data = []
    std_column_data = []

    for media in media_data:
        format = media.get('Format')
        channels = media.get('Channel(s)', '')

        if supported_formats and format not in supported_formats:
            continue
        if supported_channels and format == 'AC-3' and channels not in supported_channels:
            continue

        key_prefix = None
        if format == 'AVC':
            key_prefix = 'AVC_Video'
        elif format == 'AC-3' and channels == '2 channels':
            key_prefix = 'AC3_2_channels'
        elif format == 'AC-3' and channels == '6 channels':
            key_prefix = 'AC3_6_channels'
        elif format == 'MPEG Audio':
            key_prefix = 'MPEG_Audio'
        elif format == 'AAC LC SBR' and channels == '2 channels':
            key_prefix = 'AAC_LC_SBR_2_channels'
        elif format == 'E-AC-3' and channels == '6 channels':
            key_prefix = 'EAC3_6_channels'

        if key_prefix:
            remarks = generate_remarks(media, standard_values[key_prefix], key_prefix, keys_not_expected)
            std_values = generate_std_values(media, standard_values[key_prefix])

            remarks_column_data.append(remarks)
            std_column_data.append(std_values)

            media_json.append(flatten_json(media))
            data = list(media_json[-1].items())

            for i in range(len(std_column_data[0])):
                if remarks_column_data[0][i].startswith('Property not expected'):
                    std_column_data[0][i] = ' '

            updated_data_tuples = [(k, v, v1, remark) for (k, v), v1, remark in zip(data, std_values, remarks)]
            media_df.append(pd.DataFrame(updated_data_tuples, columns=['Item', 'Actual Value', 'Expected Value', 'Remarks']))

    return media_df

def prepare_mediaInfo_report(ts_file_path, mediaInfo_filename, ts_report_folder):
    media_info_av_data = extract_media_info(ts_file_path)
    # Initialize Lists to Store Results
    splice_clear_video_pids = []
    splice_scrambled_video_pids = []
    splice_clear_audio_pids = []
    splice_scrambled_audio_pids = []

    # Extract all Video and Audio PIDs from media_info_av_data
    video_pid_dict = {video['ID']: video for video in media_info_av_data['Video']}
    audio_pid_dict = {audio['ID']: audio for audio in media_info_av_data['Audio']}

    # Extract splice PIDs from splice_service_infos
    splice_clear_video_ids = set()
    splice_scrambled_video_ids = set()
    splice_clear_audio_ids = set()
    splice_scrambled_audio_ids = set()

    for service in splice_service_infos:
        # Video PIDs
        splice_clear_video_ids.update(service['video_pids'].get('clear', []))
        splice_scrambled_video_ids.update(service['video_pids'].get('scrambled', []))

        # Audio PIDs
        splice_clear_audio_ids.update(service['audio_pids'].get('clear', []))
        splice_scrambled_audio_ids.update(service['audio_pids'].get('scrambled', []))

    # Function to match and retrieve media info
    def retrieve_media_info(pid_ids, pid_dict, category_list):
        for pid in pid_ids:
            media_info = pid_dict.get(pid)
            if media_info:
                category_list.append(media_info)
            else:
                print(f"\n\nWarning: PID {pid} not found in media_info_av_data.")

    # Match and retrieve splice clear video PIDs
    retrieve_media_info(splice_clear_video_ids, video_pid_dict, splice_clear_video_pids)

    # Match and retrieve splice scrambled video PIDs
    retrieve_media_info(splice_scrambled_video_ids, video_pid_dict, splice_scrambled_video_pids)

    # Match and retrieve splice clear audio PIDs
    retrieve_media_info(splice_clear_audio_ids, audio_pid_dict, splice_clear_audio_pids)

    # Match and retrieve splice scrambled audio PIDs
    retrieve_media_info(splice_scrambled_audio_ids, audio_pid_dict, splice_scrambled_audio_pids)

    # Identify Non-Splice Video PIDs
    all_video_pids = set(video_pid_dict.keys())
    splice_video_pids = splice_clear_video_ids.union(splice_scrambled_video_ids)
    non_splice_video_pids = all_video_pids - splice_video_pids
    non_splice_video_pids_list = [video_pid_dict[pid] for pid in non_splice_video_pids]

    # Identify Non-Splice Audio PIDs
    all_audio_pids = set(audio_pid_dict.keys())
    splice_audio_pids = splice_clear_audio_ids.union(splice_scrambled_audio_ids)
    non_splice_audio_pids = all_audio_pids - splice_audio_pids
    non_splice_audio_pids_list = [audio_pid_dict[pid] for pid in non_splice_audio_pids]

    categorised_media_info_data = {
        "splice_clear_video_pids" : splice_clear_video_pids,
        "splice_scrambled_video_pids" : splice_scrambled_video_pids,
        "splice_clear_audio_pids" : splice_clear_audio_pids,
        "splice_scrambled_audio_pids" : splice_scrambled_audio_pids,
        "non_splice_video_pids_list" : non_splice_video_pids_list,
        "non_splice_audio_pids_list" : non_splice_audio_pids_list
    }

    #print(json.dumps(categorised_media_info_data))

    excel_path = os.path.join(ts_report_folder, mediaInfo_filename)
    write_data_to_excel(excel_path, categorised_media_info_data)

def set_column_widths(sheet):
    """
    Function to set column widths in excel sheet
    """
    min_width, max_width, fixed_width = 15, 60, 40
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        sheet.column_dimensions[column].width = adjusted_width if min_width <= max_length + 2 <= max_width else fixed_width

def write_data_to_excel(excel_path, media_info_data):
    """
    Writes data to an Excel file and formats it.
    """
    splice_clear_video_df = process_media_data(media_info_data.get('splice_clear_video_pids', []), standard_values)
    splice_clear_audio_df = process_media_data(media_info_data.get('splice_clear_audio_pids', []), standard_values, keys_not_expected, supported_audio_formats, supported_ac3_channels)
    splice_scrambled_video_df = process_media_data(media_info_data.get('splice_scrambled_video_pids', []), standard_values)
    splice_scrambled_audio_df = process_media_data(media_info_data.get('splice_scrambled_audio_pids', []), standard_values, keys_not_expected, supported_audio_formats, supported_ac3_channels)
    non_splice_video_df = process_media_data(media_info_data.get('non_splice_video_pids_list', []), standard_values)
    non_splice_audio_df = process_media_data(media_info_data.get('non_splice_audio_pids_list', []), standard_values, keys_not_expected, supported_audio_formats, supported_ac3_channels)


    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        for i, df in enumerate(splice_clear_video_df):
            df.to_excel(writer, sheet_name=f'Splice Clear Video Track {i+1}', index=False)
        for i, df in enumerate(splice_clear_audio_df):
            df.to_excel(writer, sheet_name=f'Splice Clear Audio Track {i+1}', index=False)
        for i, df in enumerate(splice_scrambled_video_df):
            df.to_excel(writer, sheet_name=f'Splice Scrambled Video Track {i+1}', index=False)
        for i, df in enumerate(splice_scrambled_audio_df):
            df.to_excel(writer, sheet_name=f'Splice Scrambled Audio Track {i+1}', index=False)
        for i, df in enumerate(non_splice_video_df):
            df.to_excel(writer, sheet_name=f'Non-Splice Video Track {i+1}', index=False)
        for i, df in enumerate(non_splice_audio_df):
            df.to_excel(writer, sheet_name=f'Non-Splice Audio Track {i+1}', index=False)

    workbook = load_workbook(excel_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        set_column_widths(sheet)
    workbook.save(excel_path)

if __name__ == "__main__":

    if len(sys.argv) <= 1:
        print("Please Pass TS_FILE_NAME as argument")
        print("Example: python stream_parsing.py \"TS_300_SCD_5_Minutes.ts\"")
        sys.exit(1)

    ts_file_basename = os.path.basename(sys.argv[1])  # Replace with the path to your transport stream file
    ts_file_name = os.path.splitext(ts_file_basename)[0]
    ts_file_path = os.path.abspath(sys.argv[1])

    scte35_report_filename = f"report_{ts_file_name}.txt"
    scte35_data_filename = "scte35_data_report.txt"
    ts_report_filename = "transport_analysis_report.txt"
    sa_report_filename = "services_analysis_report.txt"
    scte_json_filename = "scte35_parsing_report.json"
    mediaInfo_filename = f'media_info_{ts_file_name}.xlsx'

    ts_report_folder = os.path.join("report_data", ts_file_name)
    # Create the "report_data" folder if it doesn't exist
    os.makedirs(ts_report_folder, exist_ok=True)

    results_filename_data = f"""
    Results are stored in \"{ts_report_folder}\\\" directory under files :: \n
    \t    1. {scte35_report_filename}\n
    \t    2. {scte35_data_filename}\n
    \t    3. {ts_report_filename}\n
    \t    4. {sa_report_filename}\n
    \t    5. {scte_json_filename}\n
    \t    6. {mediaInfo_filename}\n
    """

    # Save the original stdout
    original_stdout = sys.stdout

    # Define the path for the ts report file within the folder
    ts_report_filename = os.path.join(ts_report_folder, ts_report_filename)
    with open(ts_report_filename, 'w') as f:
        # Redirect standard output to the file
        sys.stdout = f

        print("=============================================================================")
        print(" TS File [%s]"%(ts_file_basename))
        print("=============================================================================")

        # Run the tsp analyze command
        result = subprocess.run(['tsp', '-I', 'file', ts_file_path, '-P', 'analyze', '-O', 'drop'], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

        # Extract the TRANSPORT & SERVICES ANALYSIS REPORT
        transport_stream_report, service_analysis_report = extract_tsp_analysis_report(result.stdout)

        base_utc_time = get_base_utc_time_from_stream(transport_stream_report)

        # Print the extracted reports
        print("\n")
        print("|=============================================================================|")
        print(transport_stream_report)
        print("|=============================================================================|")

    # Define the path for the ts report file within the folder
    sa_report_filename = os.path.join(ts_report_folder, sa_report_filename)
    with open(sa_report_filename, 'w') as f:
        # Redirect standard output to the file
        sys.stdout = f

        print("=============================================================================")
        print(" TS File [%s]"%(ts_file_basename))
        print("=============================================================================")

        print("\n")
        print("|=============================================================================|")
        print(service_analysis_report)
        print("|=============================================================================|")

        # Extract splice service info from services_analysis_report.txt
        splice_service_infos = extract_service_tables_from_file(service_analysis_report)
        # Print the extracted splice info
        #for idx, splice_service_info in enumerate(splice_service_infos):
        #    print(f"Splice Info {idx+1}:")
        #    print(splice_service_info)

        # Prepare report with mediaInfo data and compare against the MPEG TS standards defined for specific projects
        prepare_mediaInfo_report(ts_file_path, mediaInfo_filename, ts_report_folder)

    # Define the path for the ts report file within the folder
    scte35_data_filename = os.path.join(ts_report_folder, scte35_data_filename)
    with open(scte35_data_filename, 'w') as f:
        # Redirect standard output to the file
        sys.stdout = f

        print("=============================================================================")
        print(" TS File [%s]"%(ts_file_basename))
        print("=============================================================================")
        print("\n")
        parse_stream_to_prepare_pat_and_pmt(ts_file_path)

        parse_pmt_section()

        scte_35_pid_list = [int(key) for key in scte_35_video_pid_mapping.keys()]
        scte_35_video_pid_list = [int(value) for value in scte_35_video_pid_mapping.values()]

        parse_stream_for_scte35_msg(ts_file_path, scte_35_pid_list, scte_35_video_pid_list)

        scte_35_pid_data = transform_list_to_dict(scte_35_pid_list)

        parse_scte35_sections()

    # Define the path for the ts report file within the folder
    scte35_report_filename = os.path.join(ts_report_folder, scte35_report_filename)
    with open(scte35_report_filename, 'w') as f:
        # Redirect standard output to the file
        sys.stdout = f

        print("=============================================================================")
        print(" TS File [%s]"%(ts_file_basename))
        print("=============================================================================")

        print("\n")
        print("=============================================================================")
        print(" SCTE-35 Analysis Report ")
        print("=============================================================================")
        print("\nSCTE-35 Service Details:")
        for index in range(len(scte_35_pid_and_pgm_num)):
            pgm_num_and_pid_value = scte_35_pid_and_pgm_num[index].split(",")
            print("    PGM_NO [%s][0x%x] SCTE-35-PID [%s][0x%x] Video PID [%d][0x%x] %s"%(pgm_num_and_pid_value[0],int(pgm_num_and_pid_value[0]),pgm_num_and_pid_value[1],int(pgm_num_and_pid_value[1]),scte_35_video_pid_list[index],int(scte_35_video_pid_list[index]),video_pid_isScrambled[scte_35_video_pid_list[index]]))

        print("\nSCTE-35 PID Occurrences:")
        for pid, count in pid_occurrences.items():
            print(f"    PID {pid}: {count} occurrences")

        print("\n------------------------------------------------")
        print("Total Unique Splice Message              : [%d] "%(len(scte_35_sections)))
        print("------------------------------------------------")
        print("Total Splice Insert Message              : [%d] "%(tot_splice_insert_msg))
        print("Total Splice Null Message                : [%d] "%(tot_splice_null_msg))
        print("Total Splice Bandwidht Reserve Message   : [%d] "%(tot_splice_bw_reserve))
        print("Total Splice Time Signal                 : [%d] "%(tot_splice_time_signal))
        print("Total Splice Unknow Types                : [%d] "%(tot_unknown_splice_type))
        print("\n")

        prepare_scte35_summary_data(ts_file_path)

    # Define the path for the ts report file within the folder
    scte_json_filename = os.path.join(ts_report_folder, scte_json_filename)
    with open(scte_json_filename, 'w') as json_file:
        # Display all consolidated scte 35 required parameters in JSON format
        json.dump(scte_35_pid_data, json_file, indent=4)

    # Define the path for the ts timeline JSON file within the folder
    stream_timeline_json_filepath = os.path.join(parent_dir, 'utils', 'stream_timeline_data.json')
    with open(stream_timeline_json_filepath, 'w') as json_file:
        # Display all consolidated scte 35 required parameters in JSON format
        json.dump(stream_timeline_data, json_file, indent=4)

    # Restore the original stdout so further prints go to the console
    sys.stdout = original_stdout

    print(results_filename_data)
